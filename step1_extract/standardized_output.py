#!/usr/bin/env python3
"""
Standardized Output Module for Step 1
Creates timestamped folders and generates standardized CSV files
"""

import json
import logging
import re
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)


def _load_category_options(rules_dir: Path) -> Tuple[List[str], List[str]]:
    """
    Load L1 and L2 category options with descriptions for Excel dropdowns.
    
    Returns:
        Tuple of (l1_options_list, l2_options_list) where each option is "CODE - Description"
    """
    l1_options = []
    l2_options = []
    
    try:
        from step1_extract.rule_loader import RuleLoader
        
        # Ensure rules_dir exists and is correct
        if not rules_dir.exists():
            # Try to find step1_rules relative to current file
            current_file = Path(__file__)
            project_root = current_file.parent.parent.parent
            rules_dir = project_root / 'step1_rules'
        
        if not rules_dir.exists():
            logger.warning(f"Rules directory not found: {rules_dir}")
            return l1_options, l2_options
        
        rule_loader = RuleLoader(rules_dir, enable_hot_reload=True)
        
        # Load L1 categories
        l1_rules = rule_loader.load_rule_file_by_name('55_categories_l1.yaml')
        if l1_rules:
            l1_categories = l1_rules.get('categories_l1', {}).get('l1_categories', [])
            for cat in l1_categories:
                cat_id = cat.get('id', '')
                cat_name = cat.get('name', '')
                if cat_id and cat_name:
                    l1_options.append(f"{cat_id} - {cat_name}")
        
        # Load L2 categories
        l2_rules = rule_loader.load_rule_file_by_name('56_categories_l2.yaml')
        if l2_rules:
            l2_categories = l2_rules.get('categories_l2', {}).get('l2_categories', [])
            for cat in l2_categories:
                cat_id = cat.get('id', '')
                cat_name = cat.get('name', '')
                if cat_id and cat_name:
                    l2_options.append(f"{cat_id} - {cat_name}")
        
        # Sort by category ID
        l1_options.sort()
        l2_options.sort()
        
        logger.debug(f"Loaded {len(l1_options)} L1 categories and {len(l2_options)} L2 categories")
        
    except Exception as e:
        logger.warning(f"Could not load category options: {e}", exc_info=True)
        # Fallback: return empty lists
        pass
    
    return l1_options, l2_options


def get_git_info() -> Tuple[str, str]:
    """Get git commit SHA and pipeline version"""
    try:
        # Get commit SHA
        result = subprocess.run(
            ['git', 'rev-parse', '--short', 'HEAD'],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent.parent
        )
        commit_sha = result.stdout.strip() if result.returncode == 0 else 'unknown'
        
        # Get tag or version
        result = subprocess.run(
            ['git', 'describe', '--tags', '--always'],
            capture_output=True,
            text=True,
            cwd=Path(__file__).parent.parent.parent
        )
        pipeline_version = result.stdout.strip() if result.returncode == 0 else 'dev'
        
        return commit_sha, pipeline_version
    except Exception as e:
        logger.warning(f"Could not get git info: {e}")
        return 'unknown', 'dev'


def parse_pack_size_uom(text: str) -> Tuple[Optional[float], Optional[float], Optional[str]]:
    """
    Parse pack_count, unit_size, unit_uom from text like "6/32 fl oz" or "2 LB X 3 CT"
    
    Returns:
        (pack_count, unit_size, unit_uom)
    """
    if not text:
        return None, None, None
    
    text = text.strip()
    
    # Pattern 1: "6/32 fl oz" or "6/32oz" → pack_count=6, unit_size=32, unit_uom=fl_oz
    pattern1 = r'(\d+)\s*/\s*(\d+(?:\.\d+)?)\s*(?:fl\s+)?(oz|fl\.?\s*oz|floz|fl_oz)'
    match = re.search(pattern1, text, re.IGNORECASE)
    if match:
        pack_count = float(match.group(1))
        unit_size = float(match.group(2))
        unit_uom = 'fl_oz'
        return pack_count, unit_size, unit_uom
    
    # Pattern 2: "2 LB X 3 CT" → pack_count=3, unit_size=2, unit_uom=lb
    pattern2 = r'(\d+(?:\.\d+)?)\s*(LB|OZ|GAL|QT|PT|KG|G|L|ML)\s*[Xx×]\s*(\d+(?:\.\d+)?)\s*(CT|COUNT|PC|PK|PKG)'
    match = re.search(pattern2, text, re.IGNORECASE)
    if match:
        unit_size = float(match.group(1))
        unit_uom_raw = match.group(2).upper()
        pack_count = float(match.group(3))
        
        # Normalize UoM
        uom_map = {
            'LB': 'lb', 'LBS': 'lb',
            'OZ': 'oz', 'OZS': 'oz',
            'GAL': 'gal', 'GALS': 'gal',
            'QT': 'qt', 'QTS': 'qt',
            'PT': 'pt', 'PTS': 'pt',
            'KG': 'kg', 'KGS': 'kg',
            'G': 'g', 'GS': 'g',
            'L': 'l', 'LS': 'l',
            'ML': 'ml', 'MLS': 'ml',
        }
        unit_uom = uom_map.get(unit_uom_raw, unit_uom_raw.lower())
        
        return pack_count, unit_size, unit_uom
    
    # Pattern 3: "32 fl oz" or "2 LB" → pack_count=None, unit_size=32, unit_uom=fl_oz
    pattern3 = r'(\d+(?:\.\d+)?)\s*(?:fl\s+)?(oz|fl\.?\s*oz|floz|fl_oz|lb|lbs|gal|qt|pt|kg|g|l|ml|ct|count|pc|pk|pkg)'
    match = re.search(pattern3, text, re.IGNORECASE)
    if match:
        unit_size = float(match.group(1))
        unit_uom_raw = match.group(2).lower()
        
        # Normalize UoM
        uom_map = {
            'oz': 'oz', 'fl.oz': 'fl_oz', 'floz': 'fl_oz', 'fl_oz': 'fl_oz',
            'lb': 'lb', 'lbs': 'lb',
            'gal': 'gal', 'gals': 'gal',
            'qt': 'qt', 'qts': 'qt',
            'pt': 'pt', 'pts': 'pt',
            'kg': 'kg', 'kgs': 'kg',
            'g': 'g', 'gs': 'g',
            'l': 'l', 'ls': 'l',
            'ml': 'ml', 'mls': 'ml',
            'ct': 'ct', 'count': 'ct', 'pc': 'ct', 'pk': 'ct', 'pkg': 'ct',
        }
        unit_uom = uom_map.get(unit_uom_raw, unit_uom_raw)
        
        return None, unit_size, unit_uom
    
    return None, None, None


def clean_canonical_name(product_name: str, item_number: Optional[str] = None, 
                         upc: Optional[str] = None) -> str:
    """
    Clean product name to create canonical_name (remove UPC/Item#/specs)
    """
    if not product_name:
        return ''
    
    # Remove UPC if present
    if upc:
        product_name = re.sub(re.escape(upc), '', product_name, flags=re.IGNORECASE)
    
    # Remove item number if present
    if item_number:
        product_name = re.sub(re.escape(item_number), '', product_name, flags=re.IGNORECASE)
    
    # Remove common size patterns (e.g., "6/32 fl oz", "2 LB", "3 CT")
    product_name = re.sub(r'\d+\s*/\s*\d+\s*(?:fl\s+)?(?:oz|fl\.?\s*oz|floz|fl_oz)', '', product_name, flags=re.IGNORECASE)
    product_name = re.sub(r'\d+(?:\.\d+)?\s*(?:LB|OZ|GAL|QT|PT|KG|G|L|ML|CT|COUNT|PC|PK|PKG)\.?', '', product_name, flags=re.IGNORECASE)
    product_name = re.sub(r'\d+(?:\.\d+)?\s*(?:X|x|×)\s*\d+\s*(?:CT|COUNT|PC|PK|PKG)', '', product_name, flags=re.IGNORECASE)
    
    # Clean up extra spaces
    product_name = re.sub(r'\s+', ' ', product_name).strip()
    
    return product_name


def transform_item_to_line(receipt_id: str, receipt_data: Dict[str, Any], 
                          item: Dict[str, Any], line_index: int) -> Dict[str, Any]:
    """
    Transform an item to match the CSV schema
    """
    # Generate line_id
    line_id = f"{receipt_id}_{line_index:04d}"
    
    # Extract fields
    product_name = item.get('product_name', '') or item.get('display_name', '') or item.get('clean_name', '')
    raw_description = product_name
    
    # Extract pack_count, unit_size, unit_uom
    # First, use the fields we've already set (from ensure_unit_size_uom_qty)
    unit_size = item.get('unit_size')
    unit_uom = item.get('unit_uom')
    pack_count = item.get('pack_count')
    
    # If not already set, try parsing from raw_uom_text or product_name
    if unit_size is None and not unit_uom:
        raw_uom_text = item.get('raw_uom_text', '') or item.get('raw_size_text', '')
        parsed_pack_count, parsed_unit_size, parsed_unit_uom = parse_pack_size_uom(raw_uom_text)
        
        # Use parsed values if we got them
        if parsed_unit_size is not None or parsed_unit_uom:
            pack_count = parsed_pack_count if pack_count is None else pack_count
            unit_size = parsed_unit_size if unit_size is None else unit_size
            unit_uom = parsed_unit_uom if not unit_uom else unit_uom
        
        # If still not found, try parsing from product_name
        if unit_size is None and not unit_uom:
            parsed_pack_count, parsed_unit_size, parsed_unit_uom = parse_pack_size_uom(product_name)
            if parsed_unit_size is not None or parsed_unit_uom:
                pack_count = parsed_pack_count if pack_count is None else pack_count
                unit_size = parsed_unit_size if unit_size is None else unit_size
                unit_uom = parsed_unit_uom if not unit_uom else unit_uom
    
    # Clean canonical_name
    item_number = str(item.get('item_number', '')).strip() or None
    upc = str(item.get('upc', '')).strip() or None
    canonical_name = clean_canonical_name(product_name, item_number, upc)
    
    # Extract transaction date
    txn_date = receipt_data.get('transaction_date') or receipt_data.get('order_date') or receipt_data.get('receipt_date', '')
    
    # Extract confidence and match reason
    confidence = item.get('category_confidence', 0.0)
    match_reason = item.get('category_source', '') or 'unknown'
    
    # Extract needs_review_reason
    needs_review_reasons = []
    if item.get('needs_review', False):
        needs_review_reasons.extend(item.get('review_reasons', []))
    if item.get('needs_category_review', False):
        needs_review_reasons.append('low_confidence' if confidence < 0.60 else 'category_review')
    if item.get('needs_quantity_review', False):
        needs_review_reasons.append('inferred_quantity')
    needs_review_reason = '; '.join(needs_review_reasons) if needs_review_reasons else ''
    
    # Extract fee_type
    fee_type = ''
    if item.get('is_fee', False):
        product_name_lower = product_name.lower()
        if 'tax' in product_name_lower:
            fee_type = 'tax'
        elif 'tip' in product_name_lower:
            fee_type = 'tip'
        elif 'bag' in product_name_lower:
            fee_type = 'bag_fee'
        elif 'service' in product_name_lower or 'fee' in product_name_lower:
            fee_type = 'service_fee'
        elif 'discount' in product_name_lower:
            fee_type = 'discount'
        else:
            fee_type = 'other_fee'
    
    # Extract UPC status and source
    upc_status = 'present' if upc else 'missing'
    upc_source = 'receipt' if upc else ''
    
    # Build line
    line = {
        'line_id': line_id,
        'source_file': receipt_data.get('filename', '') or receipt_data.get('source_file', ''),
        'vendor': receipt_data.get('vendor', '') or receipt_data.get('vendor_name', ''),
        'txn_date': txn_date,
        'raw_description': raw_description,
        'match_reason': match_reason,
        'confidence': confidence,
        'needs_review_reason': needs_review_reason,
        'canonical_name': canonical_name,
        'brand': item.get('brand', ''),
        'item_number': item_number or '',
        'upc': upc or '',
        'pack_count': pack_count if pack_count is not None else '',
        'unit_size': unit_size if unit_size is not None else '',
        'unit_uom': unit_uom or '',
        'qty': item.get('quantity', 0.0),
        'unit_price': item.get('unit_price', 0.0),
        'extended_amount': item.get('total_price', 0.0) or item.get('extended_amount', 0.0),
        'L1': item.get('l1_category', ''),
        'L2': item.get('l2_category', ''),
        'fee_type': fee_type,
        'cogs_include': not item.get('is_fee', False) and not item.get('is_summary', False),
        'upc_status': upc_status,
        'upc_source': upc_source,
    }
    
    return line


def transform_all_receipts(receipts_data: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Transform all receipts to line-level data
    """
    all_lines = []
    
    for receipt_id, receipt_data in receipts_data.items():
        items = receipt_data.get('items', [])
        
        for idx, item in enumerate(items):
            # Skip summary lines
            if item.get('is_summary', False):
                continue
            
            try:
                line = transform_item_to_line(receipt_id, receipt_data, item, idx)
                all_lines.append(line)
            except Exception as e:
                logger.warning(f"Error transforming item {idx} in receipt {receipt_id}: {e}")
                continue
    
    return all_lines


def create_standardized_output(receipts_data: Dict[str, Dict[str, Any]], 
                              output_base_dir: Path,
                              low_confidence_threshold: float = 0.60,
                              input_dir: Optional[Path] = None) -> Path:
    """
    Create standardized output in timestamped folder
    
    Returns:
        Path to the timestamped output directory
    """
    # Create timestamped folder
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_dir = output_base_dir / 'artifacts' / 'step1' / f'STEP1_{timestamp}'
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create subdirectories
    reports_dir = output_dir / 'reports'
    tables_dir = output_dir / 'tables'
    logs_dir = output_dir / 'logs'
    
    reports_dir.mkdir(parents=True, exist_ok=True)
    tables_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"Creating standardized output in: {output_dir}")
    
    # Transform all receipts to lines
    all_lines = transform_all_receipts(receipts_data)
    
    if not all_lines:
        logger.warning("No lines found to output")
        return output_dir
    
    # Create DataFrame
    df = pd.DataFrame(all_lines)
    
    # Ensure all required columns exist (fill with empty if missing)
    required_columns = [
        'line_id', 'source_file', 'vendor', 'txn_date',
        'raw_description', 'match_reason', 'confidence', 'needs_review_reason',
        'canonical_name', 'brand', 'item_number', 'upc',
        'pack_count', 'unit_size', 'unit_uom',
        'qty', 'unit_price', 'extended_amount',
        'L1', 'L2', 'fee_type', 'cogs_include',
        'upc_status', 'upc_source'
    ]
    
    for col in required_columns:
        if col not in df.columns:
            df[col] = ''
    
    # Reorder columns
    df = df[required_columns]
    
    # Write lines_step1.csv
    lines_file = tables_dir / 'lines_step1.csv'
    df.to_csv(lines_file, index=False)
    logger.info(f"Created {lines_file} with {len(df)} lines")
    
    # Create unmapped_step1.csv (L2='C99' or missing category)
    unmapped = df[(df['L2'] == 'C99') | (df['L2'] == '') | (df['L2'].isna())]
    if len(unmapped) > 0:
        unmapped_file = tables_dir / 'unmapped_step1.csv'
        unmapped.to_csv(unmapped_file, index=False)
        logger.info(f"Created {unmapped_file} with {len(unmapped)} unmapped lines")
    
    # Create low_confidence_step1.csv (confidence < threshold)
    low_confidence = df[df['confidence'] < low_confidence_threshold]
    if len(low_confidence) > 0:
        low_confidence_file = tables_dir / 'low_confidence_step1.csv'
        low_confidence.to_csv(low_confidence_file, index=False)
        logger.info(f"Created {low_confidence_file} with {len(low_confidence)} low confidence lines")
    
    # Create upc_backfill_queue.csv (rows needing UPC)
    upc_missing = df[(df['upc_status'] == 'missing') & (df['fee_type'] == '')]
    if len(upc_missing) > 0:
        upc_backfill_file = tables_dir / 'upc_backfill_queue.csv'
        upc_missing.to_csv(upc_backfill_file, index=False)
        logger.info(f"Created {upc_backfill_file} with {len(upc_missing)} lines needing UPC")
    
    # Create data_dictionary.csv
    data_dictionary = {
        'column': required_columns,
        'description': [
            'Unique line identifier (receipt_id + line_index)',
            'Source file name/path',
            'Vendor name',
            'Transaction date',
            'Raw product description from receipt',
            'Category match reason/source',
            'Category confidence score (0.0-1.0)',
            'Reason(s) why line needs review',
            'Clean product name (no UPC/Item#/specs)',
            'Product brand',
            'Item number from receipt',
            'UPC code',
            'Pack count (e.g., 6 for "6/32 fl oz")',
            'Unit size (e.g., 32 for "6/32 fl oz")',
            'Unit of measure (e.g., fl_oz, lb, ct)',
            'Quantity',
            'Unit price',
            'Extended amount (total price)',
            'L1 category code',
            'L2 category code',
            'Fee type (tax, tip, bag_fee, service_fee, discount, other_fee)',
            'Whether to include in COGS (True/False)',
            'UPC status (present/missing)',
            'UPC source (receipt/kb/etc)'
        ]
    }
    
    data_dict_df = pd.DataFrame(data_dictionary)
    data_dict_file = tables_dir / 'data_dictionary.csv'
    data_dict_df.to_csv(data_dict_file, index=False)
    logger.info(f"Created {data_dict_file}")
    
    # Generate classification report (save to reports folder)
    try:
        from .generate_classification_report import generate_classification_report
        html_path, csv_path = generate_classification_report(receipts_data, reports_dir)
        # The function already saves to reports_dir, so html_path should be correct
        logger.info(f"Created classification report: {html_path}")
    except Exception as e:
        logger.warning(f"Could not generate classification report: {e}")
    
    # Create manifest.json
    commit_sha, pipeline_version = get_git_info()
    
    # Count statistics
    total_lines = len(df)
    classified = len(df[df['L2'] != ''].dropna(subset=['L2']))
    unmapped_count = len(unmapped) if len(unmapped) > 0 else 0
    low_confidence_count = len(low_confidence) if len(low_confidence) > 0 else 0
    
    manifest = {
        'pipeline_version': pipeline_version,
        'commit_sha': commit_sha,
        'created_at': datetime.now().isoformat(),
        'thresholds': {
            'low_confidence': low_confidence_threshold
        },
        'counters': {
            'total_lines': total_lines,
            'classified': classified,
            'unmapped': unmapped_count,
            'low_confidence': low_confidence_count
        }
    }
    
    manifest_file = output_dir / 'manifest.json'
    with open(manifest_file, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2)
    logger.info(f"Created {manifest_file}")
    
    # Copy log file if it exists
    log_file = output_base_dir / 'logs' / 'step1_extract.log'
    if log_file.exists():
        dest_log = logs_dir / 'step1_run.log'
        import shutil
        shutil.copy2(log_file, dest_log)
        logger.info(f"Copied log file to {dest_log}")
    
    # Generate Excel export for human review
    try:
        excel_file = _create_excel_export(df, unmapped, low_confidence, upc_missing, tables_dir, input_dir)
        logger.info(f"Created Excel export: {excel_file}")
    except Exception as e:
        logger.warning(f"Could not create Excel export: {e}", exc_info=True)
    
    logger.info(f"✅ Standardized output complete: {output_dir}")
    
    return output_dir


def _create_excel_export(df: pd.DataFrame, unmapped: pd.DataFrame, 
                         low_confidence: pd.DataFrame, upc_missing: pd.DataFrame,
                         tables_dir: Path, input_dir: Optional[Path] = None) -> Optional[Path]:
    """Create Excel export with multiple sheets for comprehensive review"""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not available, cannot create Excel export. Install with: pip install openpyxl")
        return None
    
    excel_file = tables_dir / 'step1_review.xlsx'
    
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Sheet 1: All Lines (main data)
            df.to_excel(writer, sheet_name='All Lines', index=False)
            
            # Sheet 2: Unmapped Lines (L2='C99' or missing)
            if len(unmapped) > 0:
                unmapped.to_excel(writer, sheet_name='Unmapped', index=False)
            else:
                # Create empty sheet with same columns
                pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name='Unmapped', index=False)
            
            # Sheet 3: Low Confidence Lines
            if len(low_confidence) > 0:
                low_confidence.to_excel(writer, sheet_name='Low Confidence', index=False)
            else:
                pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name='Low Confidence', index=False)
            
            # Sheet 4: UPC Backfill Queue
            if len(upc_missing) > 0:
                upc_missing.to_excel(writer, sheet_name='UPC Needed', index=False)
            else:
                pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name='UPC Needed', index=False)
            
            # Sheet 5: Summary Statistics
            summary_data = {
                'Metric': [
                    'Total Lines',
                    'Classified Lines',
                    'Unmapped Lines',
                    'Low Confidence Lines',
                    'Lines Needing UPC',
                    'Total Receipts',
                    'Unique Vendors',
                    'Date Range'
                ],
                'Value': [
                    len(df),
                    len(df[df['L2'] != ''].dropna(subset=['L2'])),
                    len(unmapped),
                    len(low_confidence),
                    len(upc_missing),
                    df['source_file'].nunique() if 'source_file' in df.columns else 0,
                    df['vendor'].nunique() if 'vendor' in df.columns else 0,
                    f"{df['txn_date'].min() if 'txn_date' in df.columns and df['txn_date'].notna().any() else 'N/A'} to {df['txn_date'].max() if 'txn_date' in df.columns and df['txn_date'].notna().any() else 'N/A'}"
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Sheet 6: Data Dictionary
            data_dict_file = tables_dir / 'data_dictionary.csv'
            if data_dict_file.exists():
                data_dict_df = pd.read_csv(data_dict_file)
                data_dict_df.to_excel(writer, sheet_name='Data Dictionary', index=False)
            
            # Sheet 7 & 8: Category Lists (for dropdowns)
            # Load category information for dropdowns
            # Find step1_rules directory (should be at project root, same level as step1_extract)
            current_file = Path(__file__)
            # From step1_extract/standardized_output.py -> receipt_importer/step1_rules
            project_root = current_file.parent.parent  # receipt_importer/
            rules_dir = project_root / 'step1_rules'
            l1_options, l2_options = _load_category_options(rules_dir)
            
            if l1_options:
                l1_df = pd.DataFrame({'L1 Categories': l1_options})
                l1_df.to_excel(writer, sheet_name='L1_Categories', index=False)
            
            if l2_options:
                l2_df = pd.DataFrame({'L2 Categories': l2_options})
                l2_df.to_excel(writer, sheet_name='L2_Categories', index=False)
        
        # Apply Excel enhancements: dropdowns, hyperlinks, formatting
        try:
            from openpyxl import load_workbook
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(excel_file)
            
            # Check if category sheets exist
            has_l1_sheet = 'L1_Categories' in wb.sheetnames
            has_l2_sheet = 'L2_Categories' in wb.sheetnames
            
            # Apply enhancements to data sheets (All Lines, Unmapped, Low Confidence, UPC Needed)
            data_sheets = ['All Lines', 'Unmapped', 'Low Confidence', 'UPC Needed']
            
            for sheet_name in data_sheets:
                if sheet_name not in wb.sheetnames:
                    continue
                
                ws = wb[sheet_name]
                
                # Find column indices for L1, L2, and source_file
                header_row = 1
                l1_col = None
                l2_col = None
                source_file_col = None
                
                for col_idx, cell in enumerate(ws[header_row], 1):
                    if cell.value and str(cell.value).upper() == 'L1':
                        l1_col = col_idx
                    elif cell.value and str(cell.value).upper() == 'L2':
                        l2_col = col_idx
                    elif cell.value and 'source_file' in str(cell.value).lower():
                        source_file_col = col_idx
                
                # Add dropdown lists for L1 column
                if l1_col and has_l1_sheet:
                    l1_col_letter = get_column_letter(l1_col)
                    # Use INDIRECT to reference the category sheet
                    # Count rows in L1_Categories sheet
                    l1_sheet = wb['L1_Categories']
                    l1_max_row = l1_sheet.max_row
                    l1_formula = f'INDIRECT("L1_Categories!A2:A{l1_max_row}")'
                    dv = DataValidation(type="list", formula1=l1_formula, allow_blank=True)
                    dv.error = 'Invalid category. Please select from the list.'
                    dv.errorTitle = 'Invalid Entry'
                    ws.add_data_validation(dv)
                    dv.add(f"{l1_col_letter}2:{l1_col_letter}{ws.max_row}")
                    logger.debug(f"Added L1 dropdown to {sheet_name} column {l1_col_letter}")
                
                # Add dropdown lists for L2 column
                if l2_col and has_l2_sheet:
                    l2_col_letter = get_column_letter(l2_col)
                    # Use INDIRECT to reference the category sheet
                    l2_sheet = wb['L2_Categories']
                    l2_max_row = l2_sheet.max_row
                    l2_formula = f'INDIRECT("L2_Categories!A2:A{l2_max_row}")'
                    dv = DataValidation(type="list", formula1=l2_formula, allow_blank=True)
                    dv.error = 'Invalid category. Please select from the list.'
                    dv.errorTitle = 'Invalid Entry'
                    ws.add_data_validation(dv)
                    dv.add(f"{l2_col_letter}2:{l2_col_letter}{ws.max_row}")
                    logger.debug(f"Added L2 dropdown to {sheet_name} column {l2_col_letter}")
                
                # Add hyperlinks for source_file column
                if source_file_col:
                    source_file_col_letter = get_column_letter(source_file_col)
                    # Get base input directory - try multiple possible locations
                    base_input_dir = input_dir
                    if not base_input_dir or not base_input_dir.exists():
                        # Try relative to artifacts: ../../step1_input
                        artifacts_dir = tables_dir.parent
                        base_input_dir = artifacts_dir.parent.parent.parent / 'step1_input'
                        if not base_input_dir.exists():
                            # Try from project root
                            project_root = Path(__file__).parent.parent.parent
                            base_input_dir = project_root / 'data' / 'step1_input'
                    
                    hyperlink_count = 0
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f"{source_file_col_letter}{row_idx}"]
                        if cell.value:
                            source_file = str(cell.value)
                            # Try to find the file in the input directory
                            file_path = base_input_dir / source_file
                            if not file_path.exists():
                                # Try with just filename
                                file_path = base_input_dir / Path(source_file).name
                            
                            # Also try searching in subdirectories
                            if not file_path.exists():
                                for subdir in base_input_dir.iterdir():
                                    if subdir.is_dir():
                                        test_path = subdir / Path(source_file).name
                                        if test_path.exists():
                                            file_path = test_path
                                            break
                            
                            if file_path.exists():
                                try:
                                    from openpyxl.styles import Font
                                    cell.hyperlink = str(file_path.absolute())
                                    cell.font = Font(color="0000FF", underline="single")
                                    hyperlink_count += 1
                                except Exception as e:
                                    logger.debug(f"Could not create hyperlink for {source_file}: {e}")
                    
                    logger.debug(f"Added {hyperlink_count} hyperlinks for source_file in {sheet_name}")
            
            # Auto-adjust column widths for all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_file)
            logger.info("Applied Excel enhancements: dropdowns for L1/L2, hyperlinks for source_file")
        except Exception as e:
            logger.warning(f"Could not apply Excel enhancements: {e}", exc_info=True)
        
    except Exception as e:
        logger.error(f"Error creating Excel export: {e}", exc_info=True)
        return None
    
    return excel_file


def load_data_from_artifacts(artifacts_dir: Path) -> Dict[str, Any]:
    """
    Load all data from artifacts folder for report generation
    
    Args:
        artifacts_dir: Path to artifacts/step1/STEP1_YYYYMMDD_HHMM directory
        
    Returns:
        Dictionary with:
        - 'lines': DataFrame with all lines
        - 'unmapped': DataFrame with unmapped lines
        - 'low_confidence': DataFrame with low confidence lines
        - 'upc_missing': DataFrame with lines needing UPC
        - 'manifest': Dict with manifest data
        - 'receipts_data': Dict format compatible with generate_html_report
    """
    tables_dir = artifacts_dir / 'tables'
    manifest_file = artifacts_dir / 'manifest.json'
    
    data = {
        'lines': pd.DataFrame(),
        'unmapped': pd.DataFrame(),
        'low_confidence': pd.DataFrame(),
        'upc_missing': pd.DataFrame(),
        'manifest': {},
        'receipts_data': {}
    }
    
    # Load CSV files
    lines_file = tables_dir / 'lines_step1.csv'
    if lines_file.exists():
        data['lines'] = pd.read_csv(lines_file)
        logger.info(f"Loaded {len(data['lines'])} lines from {lines_file}")
    
    unmapped_file = tables_dir / 'unmapped_step1.csv'
    if unmapped_file.exists():
        data['unmapped'] = pd.read_csv(unmapped_file)
        logger.info(f"Loaded {len(data['unmapped'])} unmapped lines")
    
    low_confidence_file = tables_dir / 'low_confidence_step1.csv'
    if low_confidence_file.exists():
        data['low_confidence'] = pd.read_csv(low_confidence_file)
        logger.info(f"Loaded {len(data['low_confidence'])} low confidence lines")
    
    upc_missing_file = tables_dir / 'upc_backfill_queue.csv'
    if upc_missing_file.exists():
        data['upc_missing'] = pd.read_csv(upc_missing_file)
        logger.info(f"Loaded {len(data['upc_missing'])} lines needing UPC")
    
    # Load manifest
    if manifest_file.exists():
        with open(manifest_file, 'r', encoding='utf-8') as f:
            data['manifest'] = json.load(f)
    
    # Convert lines DataFrame to receipts_data format for compatibility with generate_html_report
    if len(data['lines']) > 0:
        data['receipts_data'] = _convert_lines_to_receipts_data(data['lines'])
    
    return data


def _convert_lines_to_receipts_data(df: pd.DataFrame) -> Dict[str, Any]:
    """Convert lines DataFrame to receipts_data format for report generation"""
    receipts_data = {}
    
    # Group by source_file to recreate receipt structure
    for source_file, group in df.groupby('source_file'):
        receipt_id = f"receipt_{source_file}"
        
        # Extract receipt metadata from first row of group
        first_row = group.iloc[0]
        
        # Convert DataFrame rows to item dictionaries
        items = []
        for _, row in group.iterrows():
            # Helper function to safely get string values from DataFrame
            def safe_get_str(col, default=''):
                val = row.get(col, default)
                if pd.isna(val) or val is None:
                    return default
                return str(val).strip() if val else default
            
            def safe_get_float(col, default=0.0):
                val = row.get(col, default)
                if pd.isna(val) or val is None:
                    return default
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return default
            
            def safe_get_bool(col, default=False):
                val = row.get(col, default)
                if pd.isna(val) or val is None:
                    return default
                if isinstance(val, bool):
                    return val
                if isinstance(val, str):
                    return val.lower() in ('true', '1', 'yes', 't')
                return bool(val)
            
            item = {
                'product_name': safe_get_str('canonical_name') or safe_get_str('raw_description', ''),
                'quantity': safe_get_float('qty', 0),
                'unit_price': safe_get_float('unit_price', 0),
                'total_price': safe_get_float('extended_amount', 0),
                'purchase_uom': safe_get_str('unit_uom', ''),
                'item_number': safe_get_str('item_number', ''),
                'upc': safe_get_str('upc', ''),
                'brand': safe_get_str('brand', ''),
                'L1': safe_get_str('L1', ''),
                'L2': safe_get_str('L2', ''),
                'category_confidence': safe_get_float('confidence', 0),
                'match_reason': safe_get_str('match_reason', ''),
                'fee_type': safe_get_str('fee_type', ''),
                'cogs_include': safe_get_bool('cogs_include', False),
            }
            items.append(item)
        
        # Helper function to safely get string from first row
        def safe_get_first_str(col, default=''):
            val = first_row.get(col, default)
            if pd.isna(val) or val is None:
                return default
            return str(val).strip() if val else default
        
        receipt_data = {
            'filename': source_file,
            'vendor': safe_get_first_str('vendor', 'Unknown'),
            'source_file': source_file,
            'transaction_date': safe_get_first_str('txn_date', ''),
            'items': items,
            'total': sum(float(item.get('total_price', 0) or 0) for item in items),
            'subtotal': sum(float(item.get('total_price', 0) or 0) for item in items if not item.get('fee_type')),
            'tax': sum(float(item.get('total_price', 0) or 0) for item in items if item.get('fee_type') == 'tax'),
            'currency': 'USD'
        }
        
        receipts_data[receipt_id] = receipt_data
    
    return receipts_data

