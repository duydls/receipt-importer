#!/usr/bin/env python3
"""
Odoo Purchase Order Excel Processor
Processes Excel files exported from Odoo containing purchase order data.

Format:
- Vendor: Vendor name (first row of each order)
- Order Lines/Description: Product name
- Order Lines/Quantity: Quantity ordered
- Order Lines/Unit Price: Price per unit
- Expected Arrival: Expected arrival date
- Order Lines/Unit of Measure: Unit of measure

Multiple products per order are indicated by empty Vendor cells.
"""

import logging
import os
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl

logger = logging.getLogger(__name__)


def process_odoo_excel(file_path: Path) -> Optional[Dict[str, Any]]:
    """
    Process an Odoo purchase order Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary containing receipt data or None if processing failed
    """
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # Extract data rows (skip header)
        data_rows = []
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (skip header)
            row_data = {}
            for col in range(1, sheet.max_column + 1):
                header_cell = sheet.cell(row=1, column=col)
                data_cell = sheet.cell(row=row, column=col)

                if header_cell.value:
                    header = str(header_cell.value).strip()
                    value = data_cell.value
                    row_data[header] = value
            data_rows.append(row_data)

        # Group rows by order (vendor name indicates start of new order)
        orders = []
        current_order = None

        for row_data in data_rows:
            vendor = row_data.get('Vendor', '').strip() if row_data.get('Vendor') else ''

            # If vendor cell is not empty, this is the start of a new order
            if vendor:
                # Save previous order if exists
                if current_order and current_order['items']:
                    orders.append(current_order)

                # Start new order
                expected_arrival = row_data.get('Expected Arrival')
                if expected_arrival:
                    # Handle different date formats
                    if isinstance(expected_arrival, datetime):
                        order_date = expected_arrival
                    elif isinstance(expected_arrival, str):
                        try:
                            order_date = datetime.strptime(expected_arrival, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            try:
                                order_date = datetime.strptime(expected_arrival, '%Y-%m-%d')
                            except ValueError:
                                order_date = datetime.now()
                    else:
                        order_date = datetime.now()
                else:
                    order_date = datetime.now()

                current_order = {
                    'vendor': vendor,
                    'order_date': order_date.isoformat(),
                    'items': []
                }

            # Add item to current order (if we have a current order)
            if current_order:
                description = row_data.get('Order Lines/Description', '').strip()
                if description:  # Only add if description exists
                    try:
                        quantity = float(row_data.get('Order Lines/Quantity', 0) or 0)
                    except (ValueError, TypeError):
                        quantity = 0

                    try:
                        unit_price = float(row_data.get('Order Lines/Unit Price', 0) or 0)
                    except (ValueError, TypeError):
                        unit_price = 0

                    uom = str(row_data.get('Order Lines/Unit of Measure', 'each')).strip()
                    if not uom:
                        uom = 'each'

                    total_price = quantity * unit_price

                    item = {
                        'product_name': description,
                        'display_name': description,
                        'canonical_name': description,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'total_price': total_price,
                        'purchase_uom': uom,
                        'is_fee': False,
                        'is_summary': False
                    }

                    current_order['items'].append(item)

        # Don't forget the last order
        if current_order and current_order['items']:
            orders.append(current_order)

        if not orders:
            logger.warning(f"No valid orders found in {file_path}")
            return None

        # Create receipt data structure
        # Use filename as receipt ID
        receipt_id = file_path.stem

        receipt_data = {
            'receipt_id': receipt_id,
            'vendor': orders[0]['vendor'] if orders else 'Unknown',
            'transaction_date': orders[0]['order_date'] if orders else datetime.now().isoformat(),
            'order_date': orders[0]['order_date'] if orders else datetime.now().isoformat(),
            'total_amount': sum(sum(item['total_price'] for item in order['items']) for order in orders),
            'items': []
        }

        # Combine all items from all orders
        for order in orders:
            receipt_data['items'].extend(order['items'])

        logger.info(f"Processed Odoo Excel file {file_path}: {len(orders)} orders, {len(receipt_data['items'])} total items")

        return {receipt_id: receipt_data}

    except Exception as e:
        logger.error(f"Error processing Odoo Excel file {file_path}: {e}", exc_info=True)
        return None


def is_odoo_excel_file(file_path: Path) -> bool:
    """
    Check if a file is an Odoo purchase order Excel file.

    Args:
        file_path: Path to the file

    Returns:
        True if this appears to be an Odoo Excel file
    """
    if not file_path.suffix.lower() == '.xlsx':
        return False

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet = wb.active

        # Check if it has the expected headers
        if sheet.max_row < 2:  # Need at least header + 1 data row
            return False

        # Get headers
        headers = []
        for col in range(1, min(7, sheet.max_column + 1)):  # Check first 6 columns
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                headers.append(str(cell.value).strip())

        # Check for expected Odoo headers
        expected_headers = [
            'Vendor',
            'Order Lines/Description',
            'Order Lines/Quantity',
            'Order Lines/Unit Price',
            'Expected Arrival',
            'Order Lines/Unit of Measure'
        ]

        # Check if we have at least some of the expected headers
        matched_headers = sum(1 for expected in expected_headers if expected in headers)
        if matched_headers >= 3:  # At least 3 matching headers
            return True

        # Also check for alternative header formats
        alt_headers = [
            'vendor',
            'description',
            'quantity',
            'unit price',
            'expected arrival',
            'unit of measure'
        ]

        alt_matched = sum(1 for alt in alt_headers if any(alt.lower() in h.lower() for h in headers))
        return alt_matched >= 3

    except Exception as e:
        logger.debug(f"Error checking if {file_path} is Odoo Excel: {e}")
        return False
