#!/usr/bin/env python3
"""
Generate SQL commands to update stock quantities and locations from Excel file
Accepts any Excel file with columns: Location, Product, Unit of Measure, Inventoried Quantity
"""

import sys
import argparse
import re
from pathlib import Path
from openpyxl import load_workbook

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from step3_mapping.query_database import connect_to_database
from psycopg2.extras import RealDictCursor


def get_product_mapping(conn):
    """Get mapping of product names to product_ids"""
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT pp.id as product_id, pt.name->>'en_US' as product_name
            FROM product_product pp
            JOIN product_template pt ON pp.product_tmpl_id = pt.id
            WHERE pt.active = true
        """)
        products = {}
        for row in cur.fetchall():
            name = row['product_name'] or ''
            if name:
                products[name.strip()] = row['product_id']
        return products


def get_location_mapping(conn):
    """Get mapping of location complete_names to location_ids"""
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT id as location_id, complete_name
            FROM stock_location
            WHERE usage = 'internal'
        """)
        locations = {}
        for row in cur.fetchall():
            name = row['complete_name'] or ''
            if name:
                locations[name.strip()] = row['location_id']
        return locations


def generate_sql_for_product(product_id, location_id, quantity, product_name="", location_name="", write_uid=2):
    """Generate SQL commands for one product with feedback"""
    sql = f"""-- Product: {product_name} (ID: {product_id}) at {location_name} (ID: {location_id}), Quantity: {quantity}

DO $$
DECLARE
    v_product_id INTEGER := {product_id};
    v_location_id INTEGER := {location_id};
    v_quantity NUMERIC := {quantity};
    v_updated INTEGER := 0;
    v_inserted INTEGER := 0;
BEGIN
    -- Step 1: Update existing record (if it exists)
    UPDATE stock_quant
    SET quantity = v_quantity,
        inventory_quantity = v_quantity,
        inventory_quantity_set = TRUE,
        inventory_date = '2025-09-01',
        in_date = '2025-09-01 00:00:00'::timestamp,
        write_uid = {write_uid},
        write_date = NOW()
    WHERE product_id = v_product_id
      AND location_id = v_location_id
      AND (lot_id IS NULL OR lot_id = 0)
      AND (package_id IS NULL OR package_id = 0)
      AND (owner_id IS NULL OR owner_id = 0);
    
    GET DIAGNOSTICS v_updated = ROW_COUNT;
    
    IF v_updated > 0 THEN
        RAISE NOTICE 'Updated % row(s) for product_id=%, location_id=%, quantity=%', v_updated, v_product_id, v_location_id, v_quantity;
    ELSE
        -- Step 2: Insert new record if UPDATE didn't affect any rows
        INSERT INTO stock_quant (
            product_id, location_id, quantity, inventory_quantity, reserved_quantity,
            inventory_quantity_set, inventory_date,
            company_id, create_uid, write_uid, create_date, write_date, in_date
        ) VALUES (
            v_product_id, v_location_id, v_quantity, v_quantity, 0,
            TRUE, '2025-09-01',
            1, {write_uid}, {write_uid}, NOW(), NOW(), '2025-09-01 00:00:00'::timestamp
        );
        
        GET DIAGNOSTICS v_inserted = ROW_COUNT;
        
        IF v_inserted > 0 THEN
            RAISE NOTICE 'Inserted % row(s) for product_id=%, location_id=%, quantity=%', v_inserted, v_product_id, v_location_id, v_quantity;
        END IF;
    END IF;
END $$;

"""
    return sql


def find_quantity_column(ws):
    """Find the column index for quantity (Inventoried Quantity)"""
    # Check header row
    headers = [cell.value for cell in ws[1]]
    
    # Look for quantity-related column names
    quantity_keywords = ['inventoried quantity', 'quantity', 'qty', 'on hand', 'stock']
    for idx, header in enumerate(headers):
        if header and any(keyword in str(header).lower() for keyword in quantity_keywords):
            return idx
    
    # Default to column D (index 3) if not found
    return 3


def main():
    """Main function"""
    parser = argparse.ArgumentParser(
        description='Generate SQL to update stock quantities from Excel file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 scripts/generate_stock_update_sql.py data/Workin_storage_template.xlsx
  python3 scripts/generate_stock_update_sql.py Freezer_template.xlsx
  python3 scripts/generate_stock_update_sql.py data/Workin_storage_template.xlsx -o custom_output.sql
        """
    )
    parser.add_argument(
        'excel_file',
        type=str,
        nargs='?',
        default=None,
        help='Path to Excel file (default: looks for Workin_storage_template.xlsx or Freezer_template.xlsx)'
    )
    parser.add_argument(
        '-o', '--output',
        type=str,
        default=None,
        help='Output SQL file path (default: based on input filename)'
    )
    parser.add_argument(
        '--write-uid',
        type=int,
        default=2,
        help='User ID for write_uid (default: 2)'
    )
    
    args = parser.parse_args()
    
    # Determine input file
    if args.excel_file:
        excel_file = Path(args.excel_file)
        if not excel_file.is_absolute():
            # Try relative to project root first
            project_root = Path(__file__).parent.parent
            if (project_root / excel_file).exists():
                excel_file = project_root / excel_file
            elif (project_root / 'data' / excel_file).exists():
                excel_file = project_root / 'data' / excel_file
    else:
        # Try default files
        project_root = Path(__file__).parent.parent
        if (project_root / 'Freezer_template.xlsx').exists():
            excel_file = project_root / 'Freezer_template.xlsx'
        elif (project_root / 'data' / 'Workin_storage_template.xlsx').exists():
            excel_file = project_root / 'data' / 'Workin_storage_template.xlsx'
        elif (project_root / 'data' / 'Walkin_storage_template.xlsx').exists():
            excel_file = project_root / 'data' / 'Walkin_storage_template.xlsx'
        else:
            print("ERROR: No Excel file specified and no default file found.")
            print("Please specify: python3 scripts/generate_stock_update_sql.py <excel_file>")
            return
    
    if not excel_file.exists():
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    # Determine output file
    if args.output:
        output_file = Path(args.output)
    else:
        # Generate output filename based on input
        output_name = excel_file.stem.replace('_template', '').replace('Template', '')
        output_file = excel_file.parent / f'update_{output_name}_stock.sql'
    
    write_uid = args.write_uid
    
    print("Connecting to database...")
    conn = connect_to_database()
    if not conn:
        print("ERROR: Could not connect to database")
        return
    
    print("Loading product mappings...")
    product_map = get_product_mapping(conn)
    print(f"  Found {len(product_map)} products")
    
    print("Loading location mappings...")
    location_map = get_location_mapping(conn)
    print(f"  Found {len(location_map)} locations")
    
    print(f"\nReading Excel file: {excel_file}")
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Find quantity column
    quantity_col_idx = find_quantity_column(ws)
    print(f"  Quantity column index: {quantity_col_idx} (0-based)")
    
    # Generate SQL
    sql_commands = []
    matched_count = 0
    not_found_products = []
    not_found_locations = []
    commented_count = 0
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0] or not row[1]:
            continue
        
        location_name = str(row[0]).strip()
        product_name = str(row[1]).strip()
        quantity = row[quantity_col_idx] if len(row) > quantity_col_idx and row[quantity_col_idx] is not None else 0
        
        # Skip if quantity is 0 or invalid
        try:
            quantity = float(quantity)
            if quantity < 0:
                continue
        except (ValueError, TypeError):
            continue
        
        # Find product_id - exact match only
        product_id = product_map.get(product_name)
        product_found = product_id is not None
        
        # Find location_id
        location_id = location_map.get(location_name)
        location_found = location_id is not None
        
        # Track issues
        if not product_found:
            not_found_products.append((row_num, product_name))
        if not location_found:
            not_found_locations.append((row_num, location_name))
        
        # Generate SQL - active if both found, commented if not
        if product_found and location_found:
            sql = generate_sql_for_product(product_id, location_id, quantity, product_name, location_name, write_uid)
            sql_commands.append(f"-- Row {row_num}: {product_name} at {location_name}\n{sql}")
            matched_count += 1
        else:
            # Generate commented SQL with placeholders
            issues = []
            if not product_found:
                issues.append(f"PRODUCT NOT FOUND: '{product_name}'")
            if not location_found:
                issues.append(f"LOCATION NOT FOUND: '{location_name}'")
            
            sql_comment = f"""-- Row {row_num}: {product_name} at {location_name}
-- ⚠️  SKIPPED: {', '.join(issues)}
-- TODO: Fix product name or location, then uncomment and update the IDs below
-- 
-- UPDATE stock_quant
-- SET quantity = {quantity},
--     inventory_quantity = {quantity},
--     write_uid = {write_uid},
--     write_date = NOW()
-- WHERE product_id = <FIX_PRODUCT_ID>  -- Find product_id for: {product_name}
--   AND location_id = <FIX_LOCATION_ID>  -- Find location_id for: {location_name}
--   AND (lot_id IS NULL OR lot_id = 0)
--   AND (package_id IS NULL OR package_id = 0)
--   AND (owner_id IS NULL OR owner_id = 0);
-- 
-- INSERT INTO stock_quant (
--     product_id, location_id, quantity, inventory_quantity, reserved_quantity,
--     company_id, create_uid, write_uid, create_date, write_date, in_date
-- )
-- SELECT <FIX_PRODUCT_ID>, <FIX_LOCATION_ID>, {quantity}, {quantity}, 0, 1, {write_uid}, {write_uid}, NOW(), NOW(), NOW()
-- WHERE NOT EXISTS (
--     SELECT 1 FROM stock_quant
--     WHERE product_id = <FIX_PRODUCT_ID>
--       AND location_id = <FIX_LOCATION_ID>
--       AND (lot_id IS NULL OR lot_id = 0)
--       AND (package_id IS NULL OR package_id = 0)
--       AND (owner_id IS NULL OR owner_id = 0)
-- );

"""
            sql_commands.append(sql_comment)
            commented_count += 1
    
    # Check if we need to delete existing records for freezer locations
    delete_existing = False
    delete_sql = ""
    if 'freezer' in excel_file.name.lower():
        # Get all freezer location IDs before closing connection
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, complete_name
                FROM stock_location
                WHERE complete_name LIKE 'WH/walkin-freezer/%'
            """)
            freezer_locations = cur.fetchall()
            if freezer_locations:
                location_ids = [str(loc['id']) for loc in freezer_locations]
                delete_existing = True
                delete_sql = f"""-- Step 1: Delete all existing stock_quant records for freezer locations
-- This ensures we start fresh and avoid incorrect matches from previous runs
DO $$
DECLARE
    v_deleted INTEGER;
BEGIN
    DELETE FROM stock_quant
    WHERE location_id IN ({', '.join(location_ids)})
      AND (lot_id IS NULL OR lot_id = 0)
      AND (package_id IS NULL OR package_id = 0)
      AND (owner_id IS NULL OR owner_id = 0);
    
    GET DIAGNOSTICS v_deleted = ROW_COUNT;
    RAISE NOTICE 'Deleted % existing stock_quant record(s) from freezer locations', v_deleted;
END $$;

"""
                print(f"  Will delete existing records from {len(freezer_locations)} freezer locations")
    
    conn.close()
    
    # Write SQL file
    print(f"\nGenerating SQL file: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("-- SQL commands to update stock quantities and locations\n")
        f.write(f"-- Generated from: {excel_file.name}\n")
        if delete_existing:
            f.write("-- NOTE: This will DELETE all existing stock_quant records for freezer locations first\n")
        f.write(f"-- Total products in Excel: {matched_count + commented_count}\n")
        f.write(f"-- Ready to execute: {matched_count} products\n")
        f.write(f"-- Needs fixing (commented): {commented_count} products\n")
        f.write(f"-- Write UID: {write_uid} (adjust if needed)\n\n")
        f.write("BEGIN;\n\n")
        if delete_existing:
            f.write(delete_sql)
        f.write("\n".join(sql_commands))
        f.write("\nCOMMIT;\n")
    
    print(f"✓ Generated SQL file: {output_file}")
    print(f"  - Total products processed: {matched_count + commented_count}")
    print(f"  - Ready to execute: {matched_count} products")
    print(f"  - Needs fixing (commented): {commented_count} products")
    print(f"  - Products not found: {len(not_found_products)}")
    print(f"  - Locations not found: {len(not_found_locations)}")
    
    if not_found_products:
        print("\n⚠ Products not found in database:")
        for row_num, name in not_found_products[:10]:  # Show first 10
            print(f"  Row {row_num}: {name}")
        if len(not_found_products) > 10:
            print(f"  ... and {len(not_found_products) - 10} more")
    
    if not_found_locations:
        print("\n⚠ Locations not found in database:")
        for row_num, name in not_found_locations[:10]:  # Show first 10
            print(f"  Row {row_num}: {name}")
        if len(not_found_locations) > 10:
            print(f"  ... and {len(not_found_locations) - 10} more")
    
    print(f"\nTo execute the SQL:")
    print(f"  psql -h uniuniuptown.shop -U your_write_user -d odoo -f {output_file}")


if __name__ == '__main__':
    main()

