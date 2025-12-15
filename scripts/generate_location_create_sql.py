#!/usr/bin/env python3
"""
Generate SQL commands to create new stock locations from Excel file
Reads Freezer_template.xlsx and generates INSERT SQL for locations that don't exist
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from step3_mapping.query_database import connect_to_database
from psycopg2.extras import RealDictCursor


def get_existing_locations(conn):
    """Get all existing locations with their complete names"""
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("""
            SELECT id, complete_name, name, location_id, parent_path
            FROM stock_location
        """)
        locations = {}
        for row in cur.fetchall():
            name = row['complete_name'] or ''
            if name:
                locations[name.strip()] = {
                    'id': row['id'],
                    'name': row['name'],
                    'location_id': row['location_id'],
                    'parent_path': row['parent_path']
                }
        return locations


def get_parent_location_info(conn, parent_path):
    """Get parent location ID and info for a given path"""
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        # Try to find parent by complete_name
        cur.execute("""
            SELECT id, location_id, parent_path, company_id, warehouse_id
            FROM stock_location
            WHERE complete_name = %s
        """, (parent_path,))
        result = cur.fetchone()
        if result:
            return {
                'id': result['id'],
                'location_id': result['location_id'],
                'parent_path': result['parent_path'],
                'company_id': result.get('company_id', 1),
                'warehouse_id': result.get('warehouse_id', 1)
            }
    return None


def parse_location_path(complete_name):
    """Parse location path like 'WH/walkin-freezer/FF01' into components"""
    parts = complete_name.split('/')
    if len(parts) < 2:
        return None
    
    # Get parent path (everything except last part)
    parent_path = '/'.join(parts[:-1])
    location_name = parts[-1]
    
    return {
        'parent_path': parent_path,
        'name': location_name,
        'full_path': complete_name
    }


def generate_location_insert_sql(location_info, parent_info, write_uid=2):
    """Generate SQL to insert a new location"""
    parent_id = parent_info['id']
    parent_parent_path = parent_info['parent_path'] or ''
    
    # Calculate parent_path: parent's parent_path + parent_id + '/'
    # If parent has no parent_path, start with just the parent_id
    if parent_parent_path:
        # parent_path format: '1/7/19/' -> child should be '1/7/19/{parent_id}/{child_id}/'
        # But we don't know child_id yet, so we'll update it after insert
        base_path = parent_parent_path
    else:
        # If parent has no parent_path, we need to build it
        # Start from root: typically '1/' for company root
        base_path = f"1/{parent_id}/"
    
    sql = f"""-- Create location: {location_info['full_path']}
DO $$
DECLARE
    v_location_id INTEGER;
    v_parent_path TEXT;
BEGIN
    -- Insert the location
    INSERT INTO stock_location (
        name, complete_name, location_id, usage, company_id, warehouse_id,
        create_uid, write_uid, create_date, write_date, active, parent_path
    )
    SELECT 
        {repr(location_info['name'])},
        {repr(location_info['full_path'])},
        {parent_id},
        'internal',
        {parent_info['company_id']},
        {parent_info['warehouse_id']},
        {write_uid},
        {write_uid},
        NOW(),
        NOW(),
        TRUE,
        NULL  -- Will be set below
    WHERE NOT EXISTS (
        SELECT 1 FROM stock_location WHERE complete_name = {repr(location_info['full_path'])}
    )
    RETURNING id INTO v_location_id;
    
    -- Update parent_path if location was inserted
    IF v_location_id IS NOT NULL THEN
        -- Build parent_path: parent's parent_path already includes parent_id, just append child_id
        -- Pattern: parent_path = '1/7/19/', child should be '1/7/19/157/'
        IF {repr(parent_parent_path)} IS NOT NULL AND {repr(parent_parent_path)} != '' THEN
            v_parent_path := {repr(parent_parent_path)} || v_location_id::TEXT || '/';
        ELSE
            -- If parent has no parent_path, build from scratch: '1/{parent_id}/{location_id}/'
            v_parent_path := '1/' || {parent_id}::TEXT || '/' || v_location_id::TEXT || '/';
        END IF;
        
        UPDATE stock_location
        SET parent_path = v_parent_path
        WHERE id = v_location_id;
        
        RAISE NOTICE 'Created location % (ID: %) with parent_path: %', {repr(location_info['full_path'])}, v_location_id, v_parent_path;
    END IF;
END $$;

"""
    return sql


def main():
    """Main function"""
    excel_file = Path(__file__).parent.parent / 'Freezer_template.xlsx'
    output_file = Path(__file__).parent.parent / 'create_freezer_locations.sql'
    
    if not excel_file.exists():
        print(f"ERROR: Excel file not found: {excel_file}")
        return
    
    print("Connecting to database...")
    conn = connect_to_database()
    if not conn:
        print("ERROR: Could not connect to database")
        return
    
    print("Loading existing locations...")
    existing_locations = get_existing_locations(conn)
    print(f"  Found {len(existing_locations)} existing locations")
    
    print(f"\nReading Excel file: {excel_file}")
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Extract unique location paths
    location_paths = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            location_path = str(row[0]).strip()
            if location_path:
                location_paths.add(location_path)
    
    print(f"  Found {len(location_paths)} unique location paths in Excel")
    
    # Parse locations and find which need to be created
    locations_to_create = []
    missing_parents = []
    
    for location_path in sorted(location_paths):
        if location_path in existing_locations:
            continue  # Already exists
        
        parsed = parse_location_path(location_path)
        if not parsed:
            print(f"  ⚠ Warning: Could not parse location path: {location_path}")
            continue
        
        # Check if parent exists
        parent_info = get_parent_location_info(conn, parsed['parent_path'])
        if not parent_info:
            missing_parents.append((location_path, parsed['parent_path']))
            continue
        
        locations_to_create.append({
            'parsed': parsed,
            'parent_info': parent_info
        })
    
    conn.close()
    
    # Generate SQL
    print(f"\nGenerating SQL file: {output_file}")
    sql_commands = []
    
    for loc_info in locations_to_create:
        sql = generate_location_insert_sql(loc_info['parsed'], loc_info['parent_info'])
        sql_commands.append(sql)
    
    # Write SQL file
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("-- SQL commands to create new stock locations\n")
        f.write(f"-- Generated from: {excel_file.name}\n")
        f.write(f"-- Total locations to create: {len(locations_to_create)}\n")
        f.write(f"-- Write UID: 2 (adjust if needed)\n\n")
        f.write("BEGIN;\n\n")
        f.write("\n".join(sql_commands))
        f.write("\nCOMMIT;\n")
    
    print(f"✓ Generated SQL file: {output_file}")
    print(f"  - Locations to create: {len(locations_to_create)}")
    print(f"  - Already exist: {len(location_paths) - len(locations_to_create) - len(missing_parents)}")
    print(f"  - Missing parents: {len(missing_parents)}")
    
    if missing_parents:
        print("\n⚠ Locations with missing parents (will be skipped):")
        for loc_path, parent_path in missing_parents[:10]:
            print(f"  {loc_path} (parent: {parent_path})")
        if len(missing_parents) > 10:
            print(f"  ... and {len(missing_parents) - 10} more")
    
    if locations_to_create:
        print(f"\nSample locations to create:")
        for loc_info in locations_to_create[:5]:
            print(f"  - {loc_info['parsed']['full_path']}")
        if len(locations_to_create) > 5:
            print(f"  ... and {len(locations_to_create) - 5} more")
    
    print(f"\nTo execute the SQL:")
    print(f"  psql -h uniuniuptown.shop -U your_write_user -d odoo -f {output_file}")


if __name__ == '__main__':
    main()

