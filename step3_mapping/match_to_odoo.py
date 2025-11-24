#!/usr/bin/env python3
"""
Match receipt items to Odoo products and UoMs
Uses the Odoo products and UoMs retrieved from the database
"""

import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from difflib import SequenceMatcher
from collections import defaultdict
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class OdooProductMatcher:
    """Match receipt items to Odoo products and UoMs"""
    
    def __init__(self, products_file: str, uoms_file: str, product_vendors_file: Optional[str] = None, mapping_file: Optional[str] = None):
        """
        Initialize matcher with Odoo products and UoMs
        
        Args:
            products_file: Path to odoo_expense_products.json
            uoms_file: Path to odoo_uoms_flat.json
            product_vendors_file: Path to odoo_product_vendors.json (optional)
            mapping_file: Path to product_standard_name_mapping.json (optional)
        """
        self.products_file = Path(products_file)
        self.uoms_file = Path(uoms_file)
        self.product_vendors_file = Path(product_vendors_file) if product_vendors_file else None
        self.mapping_file = Path(mapping_file) if mapping_file else Path('data/product_standard_name_mapping.json')
        
        # Load data
        self.products = self._load_products()
        self.uoms = self._load_uoms()
        self.product_vendors = self._load_product_vendors()
        self.product_mapping = self._load_product_mapping()
        
        # Build indexes
        self.products_index = self._build_products_index()
        self.uoms_index = self._build_uoms_index()
        self.vendor_name_map = self._build_vendor_name_map()
        
        logger.info(f"Loaded {len(self.products)} products and {len(self.uoms)} UoMs")
        if self.product_vendors:
            logger.info(f"Loaded vendor information for {len(self.product_vendors)} products")
        if self.product_mapping:
            logger.info(f"Loaded {len(self.product_mapping)} product mappings")
    
    def _load_products(self) -> List[Dict]:
        """Load Odoo products"""
        with open(self.products_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def _load_uoms(self) -> List[Dict]:
        """Load Odoo UoMs"""
        with open(self.uoms_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def _load_product_vendors(self) -> Optional[Dict]:
        """Load product-vendor relationships and prices"""
        if not self.product_vendors_file or not self.product_vendors_file.exists():
            logger.warning(f"Product vendors file not found: {self.product_vendors_file}")
            return None
        
        try:
            with open(self.product_vendors_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Convert to product_id -> vendor_info mapping
                result = {}
                for product_id_str, vendor_data in data.items():
                    product_id = int(product_id_str)
                    result[product_id] = vendor_data
                return result
        except Exception as e:
            logger.warning(f"Error loading product vendors: {e}")
            return None
    
    def _load_product_mapping(self) -> Optional[Dict]:
        """Load product standard name mappings"""
        if not self.mapping_file.exists():
            logger.debug(f"Product mapping file not found: {self.mapping_file}")
            return None
        
        try:
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"Error loading product mapping: {e}")
            return None
    
    def _build_vendor_name_map(self) -> Dict[str, List[str]]:
        """Build mapping from receipt vendor names to Odoo vendor names"""
        # Common vendor name mappings
        vendor_mappings = {
            'costco': ['Costco', 'Costco Wholesale', 'Costco Vendor Ref'],
            'amazon': ['Amazon', 'Amazon Business', 'Amazon.com'],
            'rd': ['RD', 'Restaurant Depot', 'Restaurant Depot Inc'],
            'instacart': ['Instacart', 'IC-Costco', 'IC-ALDI', 'IC-Jewel-Osco', 'IC-Mariano\'s', 'IC-Restaurant Depot'],
            'jewel': ['Jewel', 'Jewel-Osco', 'Jewel Osco'],
            'aldi': ['ALDI', 'Aldi'],
            'mariano': ['Mariano\'s', 'Marianos'],
            '88': ['88', '88 MarketPlace', '88 Marketplace'],
            'duverger': ['Duverger'],
            'foodservicedirect': ['FoodServiceDirect', 'Food Service Direct'],
            'pike': ['Pike Global Foods', 'Pike'],
            'bbi': ['BBI', 'BBI Inc', 'Bubble Tea Ingredients'],
            'uni_mousse': ['UNI_Mousse', 'UNI Mousse', 'Mousse'],
        }
        
        # Build reverse mapping: receipt vendor -> possible Odoo vendor names
        reverse_map = {}
        for key, vendor_names in vendor_mappings.items():
            for vendor_name in vendor_names:
                if vendor_name.lower() not in reverse_map:
                    reverse_map[vendor_name.lower()] = []
                reverse_map[vendor_name.lower()].extend(vendor_names)
        
        return reverse_map
    
    def _build_products_index(self) -> Dict:
        """Build searchable index of products"""
        index = {}
        
        for product in self.products:
            product_name = product.get('product_name', '').strip()
            if not product_name:
                continue
            
            product_id = product.get('product_id')
            product_name_lower = product_name.lower()
            
            # Index by exact name
            index[product_name_lower] = {
                'product_id': product_id,
                'product_name': product_name,
                'uom_name': product.get('uom_name'),
                'uom_id': product.get('uom_id'),
                'l1_id': product.get('l1_id'),
                'l1_name': product.get('l1_name'),
                'l2_id': product.get('l2_id'),
                'l2_name': product.get('l2_name'),
                'category_complete_name': product.get('category_complete_name'),
                'exact_match': True,
            }
            
            # Index by individual words (for word-based matching)
            words = [w for w in product_name_lower.split() if len(w) > 2]
            for word in words:
                if word not in index:
                    index[word] = {
                        'product_id': product_id,
                        'product_name': product_name,
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l2_id': product.get('l2_id'),
                        'exact_match': False,
                    }
        
        return index
    
    def _build_uoms_index(self) -> Dict:
        """Build searchable index of UoMs"""
        index = {}
        
        for uom in self.uoms:
            uom_name = uom.get('uom_name', '').strip()
            if not uom_name:
                continue
            
            uom_id = uom.get('uom_id')
            uom_name_lower = uom_name.lower()
            
            # Index by exact name
            index[uom_name_lower] = {
                'uom_id': uom_id,
                'uom_name': uom_name,
                'category_name': uom.get('category_name'),
                'uom_type': uom.get('uom_type'),
                'factor': uom.get('factor'),
            }
            
            # Index variations (e.g., "lb" -> "lbs", "pound")
            uom_variations = self._get_uom_variations(uom_name_lower)
            for variation in uom_variations:
                if variation not in index:
                    index[variation] = {
                        'uom_id': uom_id,
                        'uom_name': uom_name,
                        'category_name': uom.get('category_name'),
                        'uom_type': uom.get('uom_type'),
                        'factor': uom.get('factor'),
                    }
        
        return index
    
    def _get_uom_variations(self, uom_name: str) -> List[str]:
        """Get common variations of UoM names"""
        variations = []
        
        # Common mappings
        uom_mappings = {
            'unit': ['units', 'each', 'piece', 'pieces', 'pc', 'pcs'],
            'lb': ['lbs', 'pound', 'pounds', 'pound(s)'],
            'kg': ['kilogram', 'kilograms'],
            'oz': ['ounce', 'ounces', 'oz(us)', 'fl oz', 'fl oz(us)'],
            'g': ['gram', 'grams'],
            'l': ['liter', 'liters', 'litre', 'litres'],
            'ml': ['milliliter', 'milliliters', 'millilitre', 'millilitres'],
            'gal': ['gallon', 'gallons', 'gal(us)'],
            'case': ['cases'],
            'pack': ['packs', 'package', 'packages'],
            'box': ['boxes'],
            'bag': ['bags'],
        }
        
        uom_lower = uom_name.lower()
        for base, vars_list in uom_mappings.items():
            if base in uom_lower or any(v in uom_lower for v in vars_list):
                variations.extend([base] + vars_list)
                break
        
        return variations
    
    def _normalize_uom(self, uom: str) -> str:
        """Normalize UoM string for matching"""
        if not uom:
            return ''
        
        uom_lower = uom.lower().strip()
        
        # Remove common suffixes
        uom_lower = re.sub(r'\s*\(.*?\)', '', uom_lower)  # Remove (US), etc.
        uom_lower = re.sub(r'\s*ea\.?$', '', uom_lower)  # Remove "ea."
        uom_lower = re.sub(r'\s*per\s+.*$', '', uom_lower)  # Remove "per case", etc.
        
        # Common normalizations
        uom_mappings = {
            'each': 'unit',
            'pieces': 'unit',
            'piece': 'unit',
            'pcs': 'unit',
            'pc': 'unit',
            'pounds': 'lb',
            'pound': 'lb',
            'lbs': 'lb',
            'kilograms': 'kg',
            'kilogram': 'kg',
            'grams': 'g',
            'gram': 'g',
            'ounces': 'oz',
            'ounce': 'oz',
            'liters': 'l',
            'liter': 'l',
            'litres': 'l',
            'litre': 'l',
            'gallons': 'gal',
            'gallon': 'gal',
        }
        
        for key, value in uom_mappings.items():
            if key in uom_lower:
                return value
        
        return uom_lower
    
    def _normalize_product_name(self, product_name: str) -> str:
        """Normalize product name for better matching"""
        if not product_name:
            return ''
        
        name_lower = product_name.lower().strip()
        
        # Extract core product keywords FIRST (before removing descriptors)
        # This ensures we capture the product type before brand names/descriptors are removed
        
        # Extract "BBQ sauce" from product names (e.g., "PC BBQ CUP SBR 100CT" -> "bbq sauce")
        # Check for "bbq" and ("sauce" or "sbr" abbreviation)
        if 'bbq' in name_lower and ('sauce' in name_lower or 'sbr' in name_lower):
            # Extract "bbq sauce" as core product
            name_lower = 'bbq sauce'
            # Return early since we've identified the core product
            return name_lower
        
        # Extract "cheese stick" from product names (e.g., "FZ MOZZ BTTRD THCK 7LB" -> "cheese stick")
        # Check for abbreviations: "MOZZ" (mozzarella), "BTTRD" (battered), "THCK" (thick), "FZ" (frozen)
        if ('mozz' in name_lower or 'cheese' in name_lower) and ('bttrd' in name_lower or 'battered' in name_lower or 'stick' in name_lower):
            # Extract "cheese stick" as core product
            name_lower = 'cheese stick'
            return name_lower
        
        # Extract "FZ pineapple chunks" from product names (e.g., "FZ PINEAPPLE IQF 10LB" -> "FZ pineapple chunks")
        # "IQF" = Individual Quick Frozen, which are chunks
        if 'pineapple' in name_lower and 'iqf' in name_lower:
            # Extract "FZ pineapple chunks" as core product
            name_lower = 'FZ pineapple chunks'
            return name_lower
        
        # Extract "FZ strawberry chunks" from product names (e.g., "FZ STRAWBERRY WHL 10LB" -> "FZ strawberry chunks")
        # "WHL" = Whole, which are chunks (similar to IQF)
        if 'strawberry' in name_lower and 'whl' in name_lower:
            # Extract "FZ strawberry chunks" as core product
            name_lower = 'FZ strawberry chunks'
            return name_lower
        
        # Extract "candy" from product names with candy-related keywords
        # Handle abbreviations: "PB" = Peanut Butter, "SNKSZ" = Snack Size, "SLICEPOP" = Sliced Pop, "RASPB" = Raspberry
        # "LOCOCHAS" = Loco Chas (candy brand), "MIX" = Mix
        candy_keywords = ['reese', 'pb', 'snksz', 'slicepop', 'raspb', 'locochas', 'mix', 'candy', 'chocolate']
        if any(keyword in name_lower for keyword in candy_keywords) and ('snksz' in name_lower or 'slicepop' in name_lower or 'locochas' in name_lower or 'reese' in name_lower):
            # Extract "candy" as core product
            name_lower = 'candy'
            return name_lower
        
        # Extract "fry basket" from product names with fryer/basket keywords
        # Handle abbreviations: "FRYER" = Fryer, "BOILOIT" = Boil It (brand), "PCKT" = Packet/Pack
        if ('fryer' in name_lower or 'fry' in name_lower) and ('basket' in name_lower or 'boiloit' in name_lower or 'pckt' in name_lower):
            # Extract "fry basket" as core product
            name_lower = 'fry basket'
            return name_lower
        
        # Extract "chicken nuggets" from product names with chicken nugget keywords
        # Handle abbreviations: "CHX" = Chicken, "NUGGET" = Nugget, "BTRD" = Battered, "TY" = Type
        if ('chx' in name_lower or 'chicken' in name_lower or 'chix' in name_lower) and 'nugget' in name_lower:
            # Extract "chicken nuggets" as core product
            name_lower = 'chicken nuggets'
            return name_lower
        
        # Extract "chicken breast" from product names with chicken breast keywords
        # Handle abbreviations: "CHIX" = Chicken, "BREAST" = Breast, "BNLS" = Boneless, "SKLS" = Skinless
        if ('chx' in name_lower or 'chicken' in name_lower or 'chix' in name_lower) and 'breast' in name_lower:
            # Extract "chicken breast" as core product
            name_lower = 'chicken breast'
            return name_lower
        
        # Extract "whipped cream stabilizer powder" from product names
        # Handle brand names: "AERO" = brand name to remove
        if 'whipped' in name_lower and 'cream' in name_lower and ('stabilizer' in name_lower or 'stabiliser' in name_lower):
            # Extract "whipped cream stabilizer powder" as core product
            name_lower = 'whipped cream stabilizer powder'
            return name_lower
        
        # Extract "croffle dough" from product names
        # Handle brand names: "Bridor" = brand name, "Raw Butter Straight" = description
        if 'bridor' in name_lower or ('raw' in name_lower and 'butter' in name_lower and 'straight' in name_lower):
            # Extract "croffle dough" as core product
            name_lower = 'croffle dough'
            return name_lower
        
        # Extract "oreo crumbs" from product names with cookie crumbs/chocolate dirt keywords
        # Handle descriptions: "Cookie Crumbs Chocolate Dirt" = Oreo crumbs, "Crushed Cookies & Crème Crumbs" = Oreo crumbs
        if ('cookie' in name_lower and 'crumb' in name_lower) or ('chocolate' in name_lower and 'dirt' in name_lower) or ('crushed' in name_lower and 'cookie' in name_lower and 'crumb' in name_lower):
            # Extract "oreo crumbs" as core product
            name_lower = 'oreo crumbs'
            return name_lower
        
        # "trash bag" / "trash bags" -> "trash bag"
        if 'trash bag' in name_lower or 'trash bags' in name_lower:
            # Extract "trash bag" as core product
            name_lower = 'trash bag'
            return name_lower
        
        # "blow torch" -> "blow torch"
        if 'blow torch' in name_lower:
            # Extract "blow torch" as core product
            name_lower = 'blow torch'
            return name_lower
        
        # "shortening" -> "vegetable oil" (shortening is a type of vegetable oil)
        # Handle abbreviations: "SHRT" = Shortening, "CRM" = Creamy, "LQ" = Liquid
        if 'shortening' in name_lower or ('oil' in name_lower and ('shrt' in name_lower or 'shortening' in name_lower)):
            # Map shortening to vegetable oil
            name_lower = 'vegetable oil'
            return name_lower
        
        # Handle common variations
        # "almondmilk" -> "almond milk"
        name_lower = re.sub(r'almondmilk', 'almond milk', name_lower)
        name_lower = re.sub(r'coconutmilk', 'coconut milk', name_lower)
        name_lower = re.sub(r'soymilk', 'soy milk', name_lower)
        # "soy milk" should stay as "soy milk" (not "soy drink")
        
        # Remove common brand names and descriptors
        brand_patterns = [
            r'^friendly farms\s+',
            r'^kirkland signature\s+',
            r'^amazon basics\s+',
            r'^silk\s+',
            r'^pc\s+',
            r'\bpc\s+',  # PC brand name anywhere
            r'\s+original\s*$',
            r'\s+unsweetened\s*$',
            r'\s+sweetened\s*$',
            r'\s+vanilla\s*$',
            r'\s+original\s+',
            r'\s+unsweetened\s+',
            r'\s+dairy free\s*',
            r'\s+gluten free\s*',
            r'\s+vegan\s*',
            r'\s+with vitamin d.*$',
            r'\s+to help support.*$',
            r'\s+\d+\s*fl\s*oz.*$',
            r'\s+half gallon.*$',
            # Remove size/quantity descriptors (e.g., "30 Gallon", "50 count", "20% Post Consumer Recycled")
            r'\d+\s*gallon\s*',
            r'\d+\s*count\s*',
            r'\d+\s*ct\s*',  # "100CT" -> remove
            r'\d+%\s*post\s+consumer\s+recycled\s*',
            r'flextra\s+strength\s*',
            r'multipurpose\s*',
            r'drawstring\s*',
            # Remove packaging descriptors
            r'\s+cup\s*',  # "CUP" in product names
            r'\s+sbr\s*',  # "SBR" abbreviation (likely "sauce" or part of product code)
            # Remove cheese stick abbreviations
            r'^fz\s+',  # "FZ" = Frozen
            r'\s+fz\s+',  # "FZ" anywhere
            r'\s+mozz\s*',  # "MOZZ" = Mozzarella
            r'\s+bttrd\s*',  # "BTTRD" = Battered
            r'\s+thck\s*',  # "THCK" = Thick
            r'\d+\s*lb\s*',  # "7LB" = 7 pounds
            # Remove pineapple abbreviations (after extraction)
            r'\s+iqf\s*',  # "IQF" = Individual Quick Frozen (already extracted as chunks)
            r'\d+\s*lb\s*',  # "10LB" = 10 pounds
            # Remove cooking/kitchen descriptors
            r'home-grade\s*',
            r'kitchen\s+cooking\s*',
            r'with\s+lock\s*',
            r'adjustable\s+flame\s*',
            r'refillable\s*',
            r'mini\s*',
            r'lighter\s+for\s+',
            r'bbq\s*',
            r'baking\s*',
            r'brulee\s+creme\s*',
            r'light\s+cooking\s+tasks\s*',
            r'diy\s+soldering\s*',
            r'\(butane\s+gas\s+excluded\)\s*',
            # Remove shortening/oil descriptors
            r'^sunrise\s*-\s*',
            r'creamy\s+liquid\s*',
            r'\d+\s*lbs?\s*',
            r'\d+\s*pound\s*',
        ]
        
        for pattern in brand_patterns:
            name_lower = re.sub(pattern, ' ', name_lower, flags=re.IGNORECASE)
        
        # Clean up extra spaces
        name_lower = ' '.join(name_lower.split())
        
        return name_lower
    
    def _is_liquid_product(self, product_name: str) -> bool:
        """Check if product name indicates a liquid product"""
        if not product_name:
            return False
        
        name_lower = product_name.lower()
        liquid_indicators = [
            'fl oz', 'fluid ounce', 'gallon', 'half gallon', 'quart', 'pint',
            'liter', 'litre', 'ml', 'milliliter', 'l', 'gal'
        ]
        
        return any(indicator in name_lower for indicator in liquid_indicators)
    
    def _convert_uom_price(self, price: float, from_uom: str, to_uom: str) -> Optional[float]:
        """Convert price from one UoM to another using UoM factors"""
        if not price or not from_uom or not to_uom:
            return None
        
        from_uom_lower = self._normalize_uom(from_uom).lower()
        to_uom_lower = self._normalize_uom(to_uom).lower()
        
        if from_uom_lower == to_uom_lower:
            return price
        
        # Find UoM factors
        from_uom_info = None
        to_uom_info = None
        
        for uom in self.uoms:
            uom_name_lower = uom.get('uom_name', '').lower()
            if uom_name_lower == from_uom_lower or from_uom_lower in uom_name_lower:
                from_uom_info = uom
            if uom_name_lower == to_uom_lower or to_uom_lower in uom_name_lower:
                to_uom_info = uom
        
        if not from_uom_info or not to_uom_info:
            return None
        
        # Check if same category
        if from_uom_info.get('category_id') != to_uom_info.get('category_id'):
            return None
        
        # Convert using factors (assuming reference UoM has factor 1.0)
        from_factor = from_uom_info.get('factor', 1.0)
        to_factor = to_uom_info.get('factor', 1.0)
        
        # Price per reference unit
        price_per_ref = price / from_factor if from_factor else price
        # Convert to target UoM
        converted_price = price_per_ref * to_factor if to_factor else price_per_ref
        
        return converted_price
    
    def _check_vendor_match(self, receipt_vendor: str, product_id: int) -> Tuple[bool, float]:
        """Check if receipt vendor matches product vendor, return (matches, confidence_boost)"""
        if not receipt_vendor or not self.product_vendors:
            return False, 0.0
        
        receipt_vendor_lower = receipt_vendor.lower().strip()
        
        # Get product vendors
        product_vendor_info = self.product_vendors.get(product_id)
        if not product_vendor_info:
            return False, 0.0
        
        # Check vendor name mappings
        vendor_names_to_check = [receipt_vendor_lower]
        if receipt_vendor_lower in self.vendor_name_map:
            vendor_names_to_check.extend([v.lower() for v in self.vendor_name_map[receipt_vendor_lower]])
        
        # Check if any vendor matches
        for vendor_id, vendor_data in product_vendor_info.get('vendors', {}).items():
            vendor_name = vendor_data.get('vendor_name', '').lower()
            for check_name in vendor_names_to_check:
                if check_name in vendor_name or vendor_name in check_name:
                    return True, 0.2  # Boost confidence by 0.2 for vendor match
        
        return False, 0.0
    
    def _check_price_match(self, receipt_price: float, receipt_uom: str, 
                          product_id: int, product_uom: str, 
                          tolerance: float = 0.3) -> Tuple[bool, float]:
        """Check if receipt price matches product historical prices, return (matches, confidence_boost)"""
        if not receipt_price or receipt_price <= 0 or not self.product_vendors:
            return False, 0.0
        
        product_vendor_info = self.product_vendors.get(product_id)
        if not product_vendor_info:
            return False, 0.0
        
        # Collect all prices for this product
        all_prices = []
        for vendor_data in product_vendor_info.get('vendors', {}).values():
            for price_info in vendor_data.get('prices', []):
                price_unit = price_info.get('price_unit')
                price_uom = price_info.get('uom_name', '')
                if price_unit and price_uom:
                    # Convert to receipt UoM if needed
                    converted_price = self._convert_uom_price(price_unit, price_uom, receipt_uom)
                    if converted_price:
                        all_prices.append(converted_price)
                    elif price_uom.lower() == receipt_uom.lower():
                        all_prices.append(price_unit)
        
        if not all_prices:
            return False, 0.0
        
        # Check if receipt price is within tolerance of any historical price
        avg_price = sum(all_prices) / len(all_prices)
        min_price = min(all_prices)
        max_price = max(all_prices)
        
        # Check if receipt price is within tolerance
        price_diff_ratio = abs(receipt_price - avg_price) / avg_price if avg_price > 0 else 1.0
        
        if price_diff_ratio <= tolerance:
            # Calculate confidence boost based on how close the price is
            boost = 0.15 * (1.0 - price_diff_ratio / tolerance)
            return True, boost
        
        return False, 0.0
    
    def match_product(self, product_name: str, category_hint: Optional[str] = None, 
                     receipt_vendor: Optional[str] = None,
                     receipt_price: Optional[float] = None,
                     receipt_uom: Optional[str] = None,
                     receipt_id: Optional[str] = None,
                     min_similarity: float = 0.6) -> Tuple[Optional[Dict], float]:
        """
        Match receipt product name to Odoo product
        
        Args:
            product_name: Product name from receipt
            category_hint: Optional L2 category ID (e.g., "C09") to boost matching
            receipt_vendor: Optional vendor name from receipt
            receipt_price: Optional price from receipt
            receipt_uom: Optional UoM from receipt
            receipt_id: Optional receipt ID for mapping lookup
            min_similarity: Minimum similarity score (0-1)
            
        Returns:
            Tuple of (matched product dict, confidence score) or (None, 0.0)
        """
        if not product_name or not product_name.strip():
            return None, 0.0
        
        # Check mapping file first (highest priority)
        if self.product_mapping and receipt_id:
            mapping_key = f"{receipt_id}|||{product_name}"
            if mapping_key in self.product_mapping:
                mapping = self.product_mapping[mapping_key]
                product_id = mapping.get('odoo_product_id')
                
                # Find the product in our products list
                for product in self.products:
                    if product.get('product_id') == product_id:
                        match = {
                            'product_id': product.get('product_id'),
                            'product_name': product.get('product_name'),
                            'uom_name': product.get('uom_name'),
                            'uom_id': product.get('uom_id'),
                            'l1_id': product.get('l1_id'),
                            'l1_name': product.get('l1_name'),
                            'l2_id': product.get('l2_id'),
                            'l2_name': product.get('l2_name'),
                            'category_complete_name': product.get('category_complete_name'),
                        }
                        confidence = 1.0  # High confidence for manual mappings
                        logger.debug(f"Mapping file match: {product_name} (receipt: {receipt_id}) → {match['product_name']} (confidence: {confidence:.2f})")
                        return match, confidence
                
                # If product not found in products list, log warning
                logger.warning(f"Product ID {product_id} from mapping not found in products list")
        
        # Also try without receipt_id (fallback for items without receipt_id)
        if self.product_mapping:
            for key, mapping in self.product_mapping.items():
                if key.endswith(f"|||{product_name}"):
                    product_id = mapping.get('odoo_product_id')
                    for product in self.products:
                        if product.get('product_id') == product_id:
                            match = {
                                'product_id': product.get('product_id'),
                                'product_name': product.get('product_name'),
                                'uom_name': product.get('uom_name'),
                                'uom_id': product.get('uom_id'),
                                'l1_id': product.get('l1_id'),
                                'l1_name': product.get('l1_name'),
                                'l2_id': product.get('l2_id'),
                                'l2_name': product.get('l2_name'),
                                'category_complete_name': product.get('category_complete_name'),
                            }
                            confidence = 0.95  # Slightly lower confidence for fallback
                            logger.debug(f"Mapping file match (fallback): {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                            return match, confidence
        
        # Special handling: Duverger items are macarons - try to match specific flavor
        if receipt_vendor and 'duverger' in receipt_vendor.lower():
            # Normalize product name for flavor matching
            normalized_receipt = product_name.lower().strip()
            
            # Try to find a macaron product that matches the flavor
            best_match = None
            best_score = 0.0
            
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'macaron' in db_name:
                    # Try to match flavor (e.g., "Hazelnut Brittle" -> "Hazelnut" or "Brittle")
                    # Extract key words from receipt name
                    receipt_words = set(re.findall(r'\b\w+\b', normalized_receipt))
                    db_words = set(re.findall(r'\b\w+\b', db_name))
                    
                    # Remove common words
                    common_words = {'macaron', 'macarons', 'cake', 'birthday', 'the', 'a', 'an', 'and', 'or'}
                    receipt_words -= common_words
                    db_words -= common_words
                    
                    # Calculate similarity based on word overlap
                    if receipt_words and db_words:
                        overlap = len(receipt_words & db_words)
                        total_unique = len(receipt_words | db_words)
                        score = overlap / total_unique if total_unique > 0 else 0.0
                    else:
                        # Fallback: use string similarity
                        score = SequenceMatcher(None, normalized_receipt, db_name).ratio()
                    
                    if score > best_score:
                        best_score = score
                        best_match = {
                            'product_id': product.get('product_id'),
                            'product_name': product.get('product_name'),
                            'uom_name': product.get('uom_name'),
                            'uom_id': product.get('uom_id'),
                            'l1_id': product.get('l1_id'),
                            'l1_name': product.get('l1_name'),
                            'l2_id': product.get('l2_id'),
                            'l2_name': product.get('l2_name'),
                            'category_complete_name': product.get('category_complete_name'),
                        }
            
            # If we found a match, return it (even if score is low, it's still a macaron)
            if best_match:
                confidence = min(0.95, 0.7 + best_score * 0.25)  # Boost confidence for Duverger macarons
                logger.debug(f"Duverger vendor match: {product_name} (vendor: {receipt_vendor}) → {best_match['product_name']} (confidence: {confidence:.2f}, score: {best_score:.2f})")
                return best_match, confidence
        
        product_name_lower = product_name.lower().strip()
        normalized_name = self._normalize_product_name(product_name)
        
        # Check for special product patterns BEFORE exact match
        # This ensures special logic (e.g., "pineapple iqf" -> "FZ Pineapple Chunks") takes precedence
        # Special handling for FZ pineapple chunks - check BEFORE exact match
        if 'FZ pineapple chunks' in normalized_name or ('pineapple' in normalized_name and 'iqf' in product_name_lower):
            # Find "FZ Pineapple Chunks" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'pineapple' in db_name and 'chunk' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"FZ pineapple chunks special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Special handling for FZ strawberry chunks - check BEFORE exact match
        if 'FZ strawberry chunks' in normalized_name or ('strawberry' in normalized_name and 'whl' in product_name_lower):
            # Find "FZ Strawberry Chunks" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'strawberry' in db_name and 'chunk' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"FZ strawberry chunks special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Special handling for oreo crumbs - check BEFORE exact match and BEFORE candy
        # Match cookie crumbs/chocolate dirt products to "Oreo Crumbs"
        if 'oreo crumbs' in normalized_name or (('cookie' in product_name_lower and 'crumb' in product_name_lower) or ('chocolate' in product_name_lower and 'dirt' in product_name_lower) or ('crushed' in product_name_lower and 'cookie' in product_name_lower and 'crumb' in product_name_lower)):
            # Find "Oreo Crumbs" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'oreo' in db_name and 'crumb' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"Oreo crumbs special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Special handling for candy - check BEFORE exact match
        # Match candy-related products (Reese's, Sliced Pop, Loco Chas, etc.) to "Candy"
        # BUT exclude cookie crumbs/chocolate dirt products (already handled above)
        candy_keywords = ['reese', 'snksz', 'slicepop', 'locochas', 'candy']
        if 'candy' in normalized_name or any(keyword in product_name_lower for keyword in candy_keywords):
            # Skip if this is a cookie crumbs/chocolate dirt product (should be Oreo Crumbs)
            if not (('cookie' in product_name_lower and 'crumb' in product_name_lower) or ('chocolate' in product_name_lower and 'dirt' in product_name_lower) or ('crushed' in product_name_lower and 'cookie' in product_name_lower and 'crumb' in product_name_lower)):
                # Find "Candy" product
                for product in self.products:
                    db_name = product.get('product_name', '').lower()
                    if db_name == 'candy' or ('candy' in db_name and len(db_name.split()) <= 2):
                        match = {
                            'product_id': product.get('product_id'),
                            'product_name': product.get('product_name'),
                            'uom_name': product.get('uom_name'),
                            'uom_id': product.get('uom_id'),
                            'l1_id': product.get('l1_id'),
                            'l1_name': product.get('l1_name'),
                            'l2_id': product.get('l2_id'),
                            'l2_name': product.get('l2_name'),
                            'category_complete_name': product.get('category_complete_name'),
                        }
                        confidence = 0.95
                        logger.debug(f"Candy special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                        return match, confidence
        
        # Special handling for fry basket - check BEFORE exact match
        # Match fryer/basket products to "Fry Basket"
        if 'fry basket' in normalized_name or (('fryer' in product_name_lower or 'fry' in product_name_lower) and ('basket' in product_name_lower or 'boiloit' in product_name_lower or 'pckt' in product_name_lower)):
            # Find "Fry Basket" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'fry' in db_name and 'basket' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"Fry basket special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Special handling for chicken nuggets - check BEFORE exact match
        # Match chicken nugget products to "Chicken Nuggets"
        if 'chicken nuggets' in normalized_name or (('chx' in product_name_lower or 'chicken' in product_name_lower or 'chix' in product_name_lower) and 'nugget' in product_name_lower):
            # Find "Chicken Nuggets" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'chicken' in db_name and 'nugget' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"Chicken nuggets special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Special handling for chicken breast - check BEFORE exact match
        # Match chicken breast products to "Chicken Breast"
        if 'chicken breast' in normalized_name or (('chx' in product_name_lower or 'chicken' in product_name_lower or 'chix' in product_name_lower) and 'breast' in product_name_lower):
            # Find "Chicken Breast" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'chicken' in db_name and 'breast' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"Chicken breast special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Special handling for whipped cream stabilizer powder - check BEFORE exact match
        # Match whipped cream stabilizer products to "Whipped Cream Stabilizer Powder"
        if 'whipped cream stabilizer powder' in normalized_name or ('whipped' in normalized_name and 'cream' in normalized_name and 'stabilizer' in normalized_name):
            # Prefer "Powder" version if receipt mentions "powder" or has "lb" UoM
            prefer_powder = 'powder' in normalized_name or (receipt_uom and 'lb' in receipt_uom.lower())
            
            # Find matching products
            matching_products = []
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'whipped' in db_name and 'cream' in db_name and 'stabilizer' in db_name:
                    matching_products.append(product)
            
            if matching_products:
                # Prefer "Powder" version if requested
                if prefer_powder:
                    for product in matching_products:
                        db_name = product.get('product_name', '').lower()
                        if 'powder' in db_name:
                            match = {
                                'product_id': product.get('product_id'),
                                'product_name': product.get('product_name'),
                                'uom_name': product.get('uom_name'),
                                'uom_id': product.get('uom_id'),
                                'l1_id': product.get('l1_id'),
                                'l1_name': product.get('l1_name'),
                                'l2_id': product.get('l2_id'),
                                'l2_name': product.get('l2_name'),
                                'category_complete_name': product.get('category_complete_name'),
                            }
                            confidence = 0.95
                            logger.debug(f"Whipped cream stabilizer powder special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                            return match, confidence
                
                # Otherwise use first match
                product = matching_products[0]
                match = {
                    'product_id': product.get('product_id'),
                    'product_name': product.get('product_name'),
                    'uom_name': product.get('uom_name'),
                    'uom_id': product.get('uom_id'),
                    'l1_id': product.get('l1_id'),
                    'l1_name': product.get('l1_name'),
                    'l2_id': product.get('l2_id'),
                    'l2_name': product.get('l2_name'),
                    'category_complete_name': product.get('category_complete_name'),
                }
                confidence = 0.95
                logger.debug(f"Whipped cream stabilizer special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                return match, confidence
        
        # Special handling for croffle dough - check BEFORE exact match
        # Match Bridor Raw Butter Straight products to "Croffle Dough"
        if 'croffle dough' in normalized_name or ('bridor' in product_name_lower or ('raw' in product_name_lower and 'butter' in product_name_lower and 'straight' in product_name_lower)):
            # Find "Croffle Dough" product
            for product in self.products:
                db_name = product.get('product_name', '').lower()
                if 'croffle' in db_name and 'dough' in db_name:
                    match = {
                        'product_id': product.get('product_id'),
                        'product_name': product.get('product_name'),
                        'uom_name': product.get('uom_name'),
                        'uom_id': product.get('uom_id'),
                        'l1_id': product.get('l1_id'),
                        'l1_name': product.get('l1_name'),
                        'l2_id': product.get('l2_id'),
                        'l2_name': product.get('l2_name'),
                        'category_complete_name': product.get('category_complete_name'),
                    }
                    confidence = 0.95
                    logger.debug(f"Croffle dough special match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        
        # Try exact match first (original name)
        # But if multiple similar products exist (e.g., "Banana" vs "Bananas"), prefer UoM match
        if product_name_lower in self.products_index:
            match = self.products_index[product_name_lower]
            if match.get('exact_match'):
                # Check if there are multiple similar products (e.g., "banana" vs "bananas")
                # If so, prefer the one that matches UoM
                if receipt_uom:
                    # Find all products with similar names
                    similar_products = []
                    for product in self.products:
                        db_name = product.get('product_name', '').lower().strip()
                        # Check if it's a similar product (same base word)
                        if db_name == product_name_lower or db_name == normalized_name:
                            similar_products.append(product)
                        elif len(product_name_lower) > 3 and (product_name_lower in db_name or db_name in product_name_lower):
                            # Check if it's a plural/singular variant (e.g., "banana" vs "bananas")
                            base_word = product_name_lower.rstrip('s')
                            if base_word in db_name or db_name.rstrip('s') == base_word:
                                similar_products.append(product)
                    
                    # If multiple similar products, prefer UoM match
                    if len(similar_products) > 1:
                        product_uom_raw = match.get('uom_name', '')
                        product_uom = product_uom_raw.lower() if isinstance(product_uom_raw, str) else ''
                        receipt_uom_normalized = self._normalize_uom(receipt_uom).lower()
                        
                        # Check if current match's UoM matches receipt UoM
                        uom_matches = False
                        if receipt_uom_normalized == product_uom:
                            uom_matches = True
                        elif receipt_uom_normalized in ['unit', 'each'] and product_uom in ['unit', 'units', 'each']:
                            uom_matches = True
                        elif receipt_uom_normalized == 'lb' and product_uom == 'lb':
                            uom_matches = True
                        
                        # If UoM doesn't match, try to find a better match
                        if not uom_matches:
                            for product in similar_products:
                                prod_uom_raw = product.get('uom_name', '')
                                prod_uom = prod_uom_raw.lower() if isinstance(prod_uom_raw, str) else ''
                                if receipt_uom_normalized == prod_uom or \
                                   (receipt_uom_normalized in ['unit', 'each'] and prod_uom in ['unit', 'units', 'each']) or \
                                   (receipt_uom_normalized == 'lb' and prod_uom == 'lb'):
                                    # Found a better match with matching UoM
                                    match = {
                                        'product_id': product.get('product_id'),
                                        'product_name': product.get('product_name'),
                                        'uom_name': product.get('uom_name'),
                                        'uom_id': product.get('uom_id'),
                                        'l1_id': product.get('l1_id'),
                                        'l1_name': product.get('l1_name'),
                                        'l2_id': product.get('l2_id'),
                                        'l2_name': product.get('l2_name'),
                                        'category_complete_name': product.get('category_complete_name'),
                                    }
                                    logger.debug(f"Exact match with UoM preference: {product_name} (UoM: {receipt_uom}) → {match['product_name']} (UoM: {match['uom_name']})")
                                    break
                
                confidence = 1.0
                # Boost confidence if category matches
                if category_hint and match.get('l2_id') == category_hint:
                    confidence = min(1.0, confidence + 0.1)
                logger.debug(f"Exact match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                return match, confidence
        
        # Try normalized name match
        if normalized_name and normalized_name != product_name_lower:
            if normalized_name in self.products_index:
                match = self.products_index[normalized_name]
                if match.get('exact_match'):
                    # Similar UoM preference logic for normalized matches
                    if receipt_uom:
                        similar_products = []
                        for product in self.products:
                            db_name = product.get('product_name', '').lower().strip()
                            if db_name == normalized_name:
                                similar_products.append(product)
                            elif len(normalized_name) > 3:
                                base_word = normalized_name.rstrip('s')
                                if base_word in db_name or db_name.rstrip('s') == base_word:
                                    similar_products.append(product)
                        
                        if len(similar_products) > 1:
                            product_uom_raw = match.get('uom_name', '')
                            product_uom = product_uom_raw.lower() if isinstance(product_uom_raw, str) else ''
                            receipt_uom_normalized = self._normalize_uom(receipt_uom).lower()
                            
                            uom_matches = False
                            if receipt_uom_normalized == product_uom:
                                uom_matches = True
                            elif receipt_uom_normalized in ['unit', 'each'] and product_uom in ['unit', 'units', 'each']:
                                uom_matches = True
                            elif receipt_uom_normalized == 'lb' and product_uom == 'lb':
                                uom_matches = True
                            
                            if not uom_matches:
                                for product in similar_products:
                                    prod_uom_raw = product.get('uom_name', '')
                                    prod_uom = prod_uom_raw.lower() if isinstance(prod_uom_raw, str) else ''
                                    if receipt_uom_normalized == prod_uom or \
                                       (receipt_uom_normalized in ['unit', 'each'] and prod_uom in ['unit', 'units', 'each']) or \
                                       (receipt_uom_normalized == 'lb' and prod_uom == 'lb'):
                                        match = {
                                            'product_id': product.get('product_id'),
                                            'product_name': product.get('product_name'),
                                            'uom_name': product.get('uom_name'),
                                            'uom_id': product.get('uom_id'),
                                            'l1_id': product.get('l1_id'),
                                            'l1_name': product.get('l1_name'),
                                            'l2_id': product.get('l2_id'),
                                            'l2_name': product.get('l2_name'),
                                            'category_complete_name': product.get('category_complete_name'),
                                        }
                                        logger.debug(f"Normalized match with UoM preference: {product_name} (UoM: {receipt_uom}) → {match['product_name']} (UoM: {match['uom_name']})")
                                        break
                    
                    confidence = 0.95  # Slightly lower for normalized match
                    if category_hint and match.get('l2_id') == category_hint:
                        confidence = 1.0
                    # Map "Trash Liners" to "Trash bag" as standard name
                    if match.get('product_name') == 'Trash Liners' and ('trash bag' in normalized_name or 'trash bags' in normalized_name):
                        match = match.copy()
                        match['product_name'] = 'Trash bag'
                        logger.debug(f"Normalized match: {product_name} → Trash bag (mapped from Trash Liners, confidence: {confidence:.2f})")
                    else:
                        logger.debug(f"Normalized match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                    return match, confidence
        
        # Try fuzzy matching
        best_match = None
        best_score = 0.0
        
        for product in self.products:
            db_name = product.get('product_name', '').lower().strip()
            if not db_name:
                continue
            
            # Calculate similarity with original name
            score = SequenceMatcher(None, product_name_lower, db_name).ratio()
            
            # Also try normalized name
            if normalized_name and normalized_name != product_name_lower:
                normalized_score = SequenceMatcher(None, normalized_name, db_name).ratio()
                score = max(score, normalized_score)
            
            # Boost for substring match
            if product_name_lower in db_name or db_name in product_name_lower:
                score = max(score, 0.85)
            if normalized_name and (normalized_name in db_name or db_name in normalized_name):
                score = max(score, 0.9)
            
            # Boost for category match
            if category_hint and product.get('l2_id') == category_hint:
                score = min(1.0, score + 0.15)
            
            # Boost for UoM match (important when multiple similar products exist, e.g., "Banana" vs "Bananas")
            if receipt_uom:
                product_uom_raw = product.get('uom_name', '')
                product_uom = product_uom_raw.lower() if isinstance(product_uom_raw, str) else ''
                receipt_uom_normalized = self._normalize_uom(receipt_uom).lower()
                
                # Check if UoMs match (considering variations like "unit", "units", "each")
                uom_matches = False
                if receipt_uom_normalized == product_uom:
                    uom_matches = True
                elif receipt_uom_normalized in ['unit', 'each'] and product_uom in ['unit', 'units', 'each']:
                    uom_matches = True
                elif receipt_uom_normalized == 'lb' and product_uom == 'lb':
                    uom_matches = True
                
                if uom_matches:
                    score = min(1.0, score + 0.2)  # Strong boost for UoM match
                    logger.debug(f"UoM match boost: {product_name} (receipt UoM: {receipt_uom}) → {db_name} (product UoM: {product_uom})")
            
            # Boost for word overlap
            receipt_words = set(w for w in normalized_name.split() if len(w) > 2)
            db_words = set(w for w in db_name.split() if len(w) > 2)
            if receipt_words and db_words:
                word_overlap = len(receipt_words & db_words) / max(len(receipt_words), len(db_words))
                score = max(score, word_overlap * 0.9)
            
            # Special handling for trash bag products
            # "trash bag" / "trash bags" should match "Trash Liners"
            if 'trash bag' in normalized_name or 'trash bags' in normalized_name:
                if 'trash' in db_name and ('liner' in db_name or 'bag' in db_name):
                    score = max(score, 0.95)
                    logger.debug(f"Trash bag keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for shortening -> vegetable oil
            # "shortening" should match "Vegetable Oil"
            if 'shortening' in normalized_name or 'shortening' in product_name_lower:
                if 'vegetable' in db_name and 'oil' in db_name:
                    score = max(score, 0.95)
                    logger.debug(f"Shortening -> Vegetable Oil match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for cheese stick / mozzarella sticks
            # "cheese stick" / "mozz bttrd" / "mozz stx" should match "Mozzarella Sticks"
            if 'cheese stick' in normalized_name or ('mozz' in product_name_lower and ('bttrd' in product_name_lower or 'battered' in product_name_lower or 'stx' in product_name_lower)):
                if 'mozzarella' in db_name and 'stick' in db_name:
                    score = max(score, 0.95)
                    logger.debug(f"Cheese stick keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for FZ pineapple chunks
            # "FZ pineapple chunks" / "pineapple iqf" should match "FZ Pineapple Chunks"
            if 'FZ pineapple chunks' in normalized_name or ('pineapple' in normalized_name and 'iqf' in product_name_lower):
                if 'pineapple' in db_name and 'chunk' in db_name:
                    score = max(score, 0.95)
                    logger.debug(f"FZ pineapple chunks keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for FZ strawberry chunks
            # "FZ strawberry chunks" / "strawberry whl" should match "FZ Strawberry Chunks"
            if 'FZ strawberry chunks' in normalized_name or ('strawberry' in normalized_name and 'whl' in product_name_lower):
                if 'strawberry' in db_name and 'chunk' in db_name:
                    score = max(score, 0.95)
                    logger.debug(f"FZ strawberry chunks keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for oreo crumbs
            # Cookie crumbs/chocolate dirt products should match "Oreo Crumbs"
            if 'oreo crumbs' in normalized_name or (('cookie' in product_name_lower and 'crumb' in product_name_lower) or ('chocolate' in product_name_lower and 'dirt' in product_name_lower) or ('crushed' in product_name_lower and 'cookie' in product_name_lower and 'crumb' in product_name_lower)):
                if 'oreo' in db_name and 'crumb' in db_name:
                    score = max(score, 0.95)
                    logger.debug(f"Oreo crumbs keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for candy
            # Candy-related products (Reese's, Sliced Pop, Loco Chas, etc.) should match "Candy"
            # BUT exclude cookie crumbs/chocolate dirt products (already handled above)
            candy_keywords = ['reese', 'snksz', 'slicepop', 'locochas', 'candy']
            if 'candy' in normalized_name or any(keyword in product_name_lower for keyword in candy_keywords):
                # Skip if this is a cookie crumbs/chocolate dirt product (should be Oreo Crumbs)
                if not (('cookie' in product_name_lower and 'crumb' in product_name_lower) or ('chocolate' in product_name_lower and 'dirt' in product_name_lower) or ('crushed' in product_name_lower and 'cookie' in product_name_lower and 'crumb' in product_name_lower)):
                    if db_name == 'candy' or ('candy' in db_name and len(db_name.split()) <= 2):
                        score = max(score, 0.95)
                        logger.debug(f"Candy keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for fry basket
            # Fryer/basket products should match "Fry Basket"
            if 'fry basket' in normalized_name or (('fryer' in product_name_lower or 'fry' in product_name_lower) and ('basket' in product_name_lower or 'boiloit' in product_name_lower or 'pckt' in product_name_lower)):
                if 'fry' in db_name and 'basket' in db_name:
                    score = max(score, 0.95)
                    logger.debug(f"Fry basket keyword match: {product_name} → {db_name} (score: {score:.2f})")
            
            # Special handling for milk products: if receipt contains "almond milk" keywords and DB has "almond milk"
            if 'almond' in normalized_name and 'milk' in normalized_name:
                if 'almond' in db_name and 'milk' in db_name:
                    score = max(score, 0.9)
                    # Penalize powder matches for liquid products
                    if 'powder' in db_name and self._is_liquid_product(product_name):
                        score = max(0.0, score - 0.5)
            if 'coconut' in normalized_name and 'milk' in normalized_name:
                if 'coconut' in db_name and 'milk' in db_name:
                    score = max(score, 0.9)
                    # Penalize powder matches for liquid products
                    if 'powder' in db_name and self._is_liquid_product(product_name):
                        score = max(0.0, score - 0.5)
            if 'soy' in normalized_name and 'milk' in normalized_name:
                if 'soy' in db_name and 'milk' in db_name:
                    score = max(score, 0.9)
                    # Penalize powder matches for liquid products
                    if 'powder' in db_name and self._is_liquid_product(product_name):
                        score = max(0.0, score - 0.5)
                # Also match "soymilk" (one word) to "soy milk"
                if db_name == 'soymilk' and 'soy' in normalized_name and 'milk' in normalized_name:
                    score = max(score, 0.95)
                    # Boost if receipt indicates liquid
                    if self._is_liquid_product(product_name):
                        score = max(score, 0.98)
            
            # Boost for vendor match
            if receipt_vendor:
                vendor_matches, vendor_boost = self._check_vendor_match(receipt_vendor, product.get('product_id'))
                if vendor_matches:
                    score = min(1.0, score + vendor_boost)
            
            # Boost for price match (if we have price and UoM info)
            if receipt_price and receipt_uom:
                price_matches, price_boost = self._check_price_match(
                    receipt_price, receipt_uom,
                    product.get('product_id'),
                    product.get('uom_name', '')
                )
                if price_matches:
                    score = min(1.0, score + price_boost)
            
            if score > best_score:
                best_score = score
                best_match = {
                    'product_id': product.get('product_id'),
                    'product_name': product.get('product_name'),
                    'uom_name': product.get('uom_name'),
                    'uom_id': product.get('uom_id'),
                    'l1_id': product.get('l1_id'),
                    'l1_name': product.get('l1_name'),
                    'l2_id': product.get('l2_id'),
                    'l2_name': product.get('l2_name'),
                    'category_complete_name': product.get('category_complete_name'),
                }
        
        # Check if best match meets threshold
        if best_match and best_score >= min_similarity:
            # Map "Trash Liners" to "Trash bag" as standard name
            if best_match.get('product_name') == 'Trash Liners' and ('trash bag' in normalized_name or 'trash bags' in normalized_name):
                best_match = best_match.copy()
                best_match['product_name'] = 'Trash bag'
                logger.debug(f"Fuzzy match: {product_name} → Trash bag (mapped from Trash Liners, score: {best_score:.2f})")
            else:
                logger.debug(f"Fuzzy match: {product_name} → {best_match['product_name']} (score: {best_score:.2f})")
            return best_match, best_score
        
        # Try word-based matching with normalized name
        receipt_words = set(w for w in normalized_name.split() if len(w) > 2)
        for word in receipt_words:
            if word in self.products_index:
                match = self.products_index[word]
                confidence = 0.7  # Lower confidence for word-based match
                if category_hint and match.get('l2_id') == category_hint:
                    confidence = 0.8
                logger.debug(f"Word-based match: {product_name} → {match['product_name']} (confidence: {confidence:.2f})")
                return match, confidence
        
        return None, 0.0
    
    def match_uom(self, purchase_uom: str, product_uom_hint: Optional[str] = None) -> Tuple[Optional[Dict], float]:
        """
        Match receipt UoM to Odoo UoM
        
        Args:
            purchase_uom: UoM from receipt
            product_uom_hint: Optional UoM from matched product (to boost matching)
            
        Returns:
            Tuple of (matched UoM dict, confidence score) or (None, 0.0)
        """
        if not purchase_uom:
            return None, 0.0
        
        purchase_uom_normalized = self._normalize_uom(purchase_uom)
        
        # Try exact match
        if purchase_uom_normalized in self.uoms_index:
            match = self.uoms_index[purchase_uom_normalized]
            confidence = 1.0
            # Boost if matches product UoM
            if product_uom_hint and isinstance(product_uom_hint, str) and match['uom_name'].lower() == product_uom_hint.lower():
                confidence = 1.0
            return match, confidence
        
        # Try direct match with original
        purchase_uom_lower = purchase_uom.lower().strip()
        if purchase_uom_lower in self.uoms_index:
            match = self.uoms_index[purchase_uom_lower]
            confidence = 0.95
            if product_uom_hint and isinstance(product_uom_hint, str) and match['uom_name'].lower() == product_uom_hint.lower():
                confidence = 1.0
            return match, confidence
        
        # Try partial match
        best_match = None
        best_score = 0.0
        
        for uom in self.uoms:
            uom_name = uom.get('uom_name', '').lower().strip()
            if not uom_name:
                continue
            
            # Check if purchase UoM contains UoM name or vice versa
            if purchase_uom_lower in uom_name or uom_name in purchase_uom_lower:
                score = 0.8
                # Boost if matches product UoM
                if product_uom_hint and isinstance(product_uom_hint, str) and uom_name == product_uom_hint.lower():
                    score = 0.95
                
                if score > best_score:
                    best_score = score
                    best_match = {
                        'uom_id': uom.get('uom_id'),
                        'uom_name': uom.get('uom_name'),
                        'category_name': uom.get('category_name'),
                        'uom_type': uom.get('uom_type'),
                        'factor': uom.get('factor'),
                    }
        
        if best_match:
            return best_match, best_score
        
        return None, 0.0
    
    def match_receipt_item(self, item: Dict) -> Dict:
        """
        Match a single receipt item to Odoo product and UoM
        
        Args:
            item: Receipt item dict
            
        Returns:
            Dict with matching results
        """
        product_name = item.get('product_name') or item.get('display_name') or item.get('canonical_name') or ''
        purchase_uom = item.get('purchase_uom') or item.get('unit_uom') or ''
        category_hint = item.get('l2_category')  # e.g., "C09"
        receipt_vendor = item.get('receipt_vendor') or item.get('vendor') or ''
        receipt_price = item.get('unit_price') or item.get('price')
        receipt_quantity = item.get('quantity', 1)
        
        # Match product (with vendor and price hints)
        product_match, product_confidence = self.match_product(
            product_name, 
            category_hint=category_hint,
            receipt_vendor=receipt_vendor,
            receipt_price=receipt_price,
            receipt_uom=purchase_uom,
            receipt_id=item.get('receipt_id')
        )
        
        # Match UoM (use product's UoM as hint if available)
        product_uom_hint_raw = product_match.get('uom_name') if product_match else None
        product_uom_hint = product_uom_hint_raw if isinstance(product_uom_hint_raw, str) else None
        uom_match, uom_confidence = self.match_uom(purchase_uom, product_uom_hint)
        
        # Determine overall match status
        matched = product_match is not None and uom_match is not None
        overall_confidence = (product_confidence + uom_confidence) / 2 if matched else max(product_confidence, uom_confidence)
        
        # Determine if review is needed
        needs_review = not matched or overall_confidence < 0.8
        
        return {
            'receipt_item': item,
            'product_match': product_match,
            'product_confidence': product_confidence,
            'uom_match': uom_match,
            'uom_confidence': uom_confidence,
            'matched': matched,
            'overall_confidence': overall_confidence,
            'needs_review': needs_review,
        }


def load_all_receipt_items(step1_output_dir: Path) -> List[Dict]:
    """Load all receipt items from Step 1 output"""
    all_items = []
    
    # Find all extracted_data.json files
    for json_file in step1_output_dir.rglob('extracted_data.json'):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                receipts = json.load(f)
            
            for receipt_id, receipt_data in receipts.items():
                vendor = receipt_data.get('vendor', 'Unknown')
                transaction_date = receipt_data.get('transaction_date', '')
                
                for item in receipt_data.get('items', []):
                    item['receipt_id'] = receipt_id
                    item['receipt_vendor'] = vendor
                    item['receipt_date'] = transaction_date
                    all_items.append(item)
        except Exception as e:
            logger.warning(f"Error loading {json_file}: {e}")
    
    return all_items


def generate_matching_report(matched_items: List[Dict], output_file: Path):
    """Generate Excel report of matching results"""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    # Prepare data for Excel
    report_data = []
    
    for match_result in matched_items:
        item = match_result['receipt_item']
        product_match = match_result.get('product_match')
        uom_match = match_result.get('uom_match')
        
        report_data.append({
            'Receipt ID': item.get('receipt_id', ''),
            'Vendor': item.get('receipt_vendor', ''),
            'Date': item.get('receipt_date', ''),
            'Receipt Product Name': item.get('product_name', ''),
            'Receipt UoM': item.get('purchase_uom', ''),
            'Receipt Quantity': item.get('quantity', ''),
            'Receipt Unit Price': item.get('unit_price', ''),
            'Receipt Total Price': item.get('total_price', ''),
            'Receipt L2 Category': item.get('l2_category', ''),
            'Matched': 'Yes' if match_result.get('matched') else 'No',
            'Overall Confidence': f"{match_result.get('overall_confidence', 0):.2f}",
            'Odoo Product ID': product_match.get('product_id') if product_match else '',
            'Odoo Product Name': product_match.get('product_name') if product_match else '',
            'Odoo L2 Category': product_match.get('l2_id') if product_match else '',
            'Odoo UoM ID': uom_match.get('uom_id') if uom_match else '',
            'Odoo UoM Name': uom_match.get('uom_name') if uom_match else '',
            'Product Confidence': f"{match_result.get('product_confidence', 0):.2f}",
            'UoM Confidence': f"{match_result.get('uom_confidence', 0):.2f}",
            'Needs Review': 'Yes' if match_result.get('needs_review') else 'No',
        })
    
    df = pd.DataFrame(report_data)
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Sheet 1: All items
        df.to_excel(writer, sheet_name='All Items', index=False)
        
        # Sheet 2: Needs review
        needs_review_df = df[df['Needs Review'] == 'Yes'].copy()
        needs_review_df.to_excel(writer, sheet_name='Needs Review', index=False)
        
        # Sheet 3: Matched items
        matched_df = df[df['Matched'] == 'Yes'].copy()
        matched_df.to_excel(writer, sheet_name='Matched', index=False)
        
        # Sheet 4: Unmatched items
        unmatched_df = df[df['Matched'] == 'No'].copy()
        unmatched_df.to_excel(writer, sheet_name='Unmatched', index=False)
        
        # Format all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width


if __name__ == '__main__':
    import sys
    
    # Paths
    step1_output_dir = Path('data/step1_output')
    products_file = 'data/odoo_expense_products.json'
    uoms_file = 'data/odoo_uoms_flat.json'
    product_vendors_file = 'data/odoo_product_vendors.json'
    output_file = Path('data/product_matching_report.xlsx')
    
    # Load receipt items
    print("Loading receipt items from Step 1 output...")
    receipt_items = load_all_receipt_items(step1_output_dir)
    print(f"✓ Loaded {len(receipt_items)} receipt items")
    
    # Initialize matcher
    print("\nInitializing Odoo product matcher...")
    mapping_file = 'data/product_standard_name_mapping.json'
    matcher = OdooProductMatcher(products_file, uoms_file, product_vendors_file, mapping_file)
    
    # Match all items
    print(f"\nMatching {len(receipt_items)} items to Odoo products...")
    matched_items = []
    for i, item in enumerate(receipt_items, 1):
        if i % 100 == 0:
            print(f"  Processed {i}/{len(receipt_items)} items...")
        match_result = matcher.match_receipt_item(item)
        matched_items.append(match_result)
    
    # Generate statistics
    matched_count = sum(1 for m in matched_items if m.get('matched'))
    needs_review_count = sum(1 for m in matched_items if m.get('needs_review'))
    high_confidence_count = sum(1 for m in matched_items if m.get('matched') and m.get('overall_confidence', 0) >= 0.8)
    
    print(f"\n{'='*80}")
    print("Matching Results:")
    print(f"{'='*80}")
    print(f"Total items: {len(matched_items)}")
    print(f"Matched: {matched_count} ({matched_count/len(matched_items)*100:.1f}%)")
    print(f"High confidence (≥0.8): {high_confidence_count} ({high_confidence_count/len(matched_items)*100:.1f}%)")
    print(f"Needs review: {needs_review_count} ({needs_review_count/len(matched_items)*100:.1f}%)")
    print(f"Unmatched: {len(matched_items) - matched_count} ({(len(matched_items) - matched_count)/len(matched_items)*100:.1f}%)")
    
    # Generate report
    print(f"\nGenerating Excel report...")
    generate_matching_report(matched_items, output_file)
    print(f"✓ Report saved to: {output_file}")
    
    # Save JSON for programmatic access
    json_output_file = Path('data/product_matching_results.json')
    with open(json_output_file, 'w', encoding='utf-8') as f:
        json.dump(matched_items, f, indent=2, ensure_ascii=False, default=str)
    print(f"✓ JSON results saved to: {json_output_file}")

